#!/usr/bin/env python3
"""
Verificación del Weekly Report usando datos y FX rates del template manual.

Lee el DATA directamente de Book 1.xlsx (datos manuales) en vez de generarlo
desde los CSVs, y usa los FX rates exactos del template. Así se puede comparar
1:1 con el Weekly hecho a mano.

Uso:
    python3 verify_weekly.py --week 8
    python3 verify_weekly.py --week 8 --entity mexico
"""

import argparse
import os
import sys
from datetime import datetime

from openpyxl import load_workbook as _load_wb

# ── FX rates EXACTOS del template (DATA SERV, AM-AO) ───────────────────
TEMPLATE_FX = {
    "EUR": {"EUR": 1.0, "USD": 1.16},
    "USD": {"EUR": 0.86, "USD": 1.0},
    "CHF": {"EUR": 1.07, "USD": 1.25},
    "GBP": {"EUR": 1.15, "USD": 1.34},
    "GPB": {"EUR": 1.15, "USD": 1.34},
    "JPY": {"EUR": 0.0055, "USD": 0.0064},
    "MXN": {"EUR": 0.048, "USD": 0.056},
}

# ── Monkey-patch: forzar rates de template para DATA SERV ───────────────
import src.fx_rates as fx_mod


def _get_template_fx(date=None):
    return {k: dict(v) for k, v in TEMPLATE_FX.items()}


fx_mod.get_historical_fx = _get_template_fx

# ── Imports ─────────────────────────────────────────────────────────────
from config import ENTITIES, TEMPLATE_PATH, BASE_DIR
from src.data_loader import load_reserva, load_dreserva, compute_rentabilidad
from src.fx_rates import preload_fx_months
from src.validators import detect_errors
from src.weekly import (
    build_serv_rows, generate_weekly_excel,
    write_diferencia_pct, write_errores_sheet,
    update_pivot_week_filters,
)
from src.bookings import build_booking_matrix, write_booking_to_excel

# Ruta del archivo de datos manual
BOOK1_PATH = os.path.join(BASE_DIR, "data", "Book 1.xlsx")

# Mapeo columna Excel → letra clave (igual que build_data_rows)
_COL_TO_KEY = {
    1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F", 7: "G",
    8: "H", 9: "I", 10: "J", 11: "K", 12: "L", 13: "M", 14: "N",
    15: "O", 16: "P", 17: "Q", 18: "R", 19: "S", 20: "T", 21: "U",
    22: "V", 23: "W", 24: "X", 26: "Z",
}


def _parse_date(val):
    """Convierte un valor a datetime si es string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        import pandas as pd
        return pd.to_datetime(val)
    except Exception:
        return None


# Columnas que contienen fechas
_DATE_COLS = {"D", "E", "F", "V"}


def _load_data_from_book1(path):
    """Lee la hoja DATA de Book 1.xlsx y retorna list of dicts (mismo formato que build_data_rows)."""
    print(f"  Leyendo DATA desde {os.path.basename(path)}...")
    wb = _load_wb(path, data_only=True)
    ws = wb["DATA"]

    rows = []
    for r in range(2, ws.max_row + 1):
        folio = ws.cell(r, 2).value
        if folio is None:
            continue

        row = {}
        for col_num, key in _COL_TO_KEY.items():
            val = ws.cell(r, col_num).value
            if key in _DATE_COLS:
                val = _parse_date(val)
            row[key] = val
        rows.append(row)

    wb.close()
    print(f"  {len(rows)} filas leídas de Book 1.xlsx")
    return rows


# ── Filtros de pivots (usa módulo compartido) ─────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Genera Weekly Report con datos de Book 1.xlsx y FX rates de template"
    )
    parser.add_argument("--week", type=int, default=None,
                        help="Número de semana (default: semana actual)")
    parser.add_argument("--entity", type=str, default="all",
                        choices=["espana", "mexico", "all"],
                        help="Entidad a procesar (default: all)")
    args = parser.parse_args()

    today = datetime.now()
    week_num = args.week or int(today.strftime('%W'))

    print(f"═══ VERIFICACIÓN Weekly Report — Semana {week_num} ═══")
    print(f"  FX rates exactos del template")
    print(f"  EUR→USD: {TEMPLATE_FX['EUR']['USD']}, USD→EUR: {TEMPLATE_FX['USD']['EUR']}")
    print(f"  MXN→EUR: {TEMPLATE_FX['MXN']['EUR']}, MXN→USD: {TEMPLATE_FX['MXN']['USD']}")

    # ── Entidades ────────────────────────────────────────────────────────
    if args.entity == "all":
        entities = ENTITIES
    else:
        entities = {args.entity: ENTITIES[args.entity]}

    # ── DATA: leer directamente de Book 1.xlsx ───────────────────────────
    if not os.path.isfile(BOOK1_PATH):
        print(f"\nERROR: No se encuentra {BOOK1_PATH}")
        sys.exit(1)

    all_data_rows = _load_data_from_book1(BOOK1_PATH)

    # Filtrar por entidad si se especificó
    if args.entity != "all":
        company = entities[args.entity]["company"]
        before = len(all_data_rows)
        all_data_rows = [r for r in all_data_rows if r.get("A") == company]
        print(f"  Filtro entidad {company}: {before} → {len(all_data_rows)} filas")

    if not all_data_rows:
        print("\nERROR: No se cargaron datos de Book 1.xlsx")
        sys.exit(1)

    # ── DATA SERV: generar desde CSVs (misma lógica que main.py) ─────────
    all_serv_rows = []
    all_errors = []
    total_serv_count = 0

    for key, cfg in entities.items():
        company = cfg["company"]
        label = cfg["label"]
        data_dir = cfg["data_dir"]

        if not os.path.exists(data_dir):
            print(f"  Saltando {label}: {data_dir} no encontrado")
            continue

        print(f"\n  Cargando DATA SERV de {label} ({company})...")
        reserva_df = load_reserva(data_dir)
        dreserva_df = load_dreserva(data_dir)

        preload_fx_months(reserva_df["fecha"].dropna())

        serv_rows, serv_count = build_serv_rows(dreserva_df, reserva_df, company)

        rentabilidad_df = compute_rentabilidad(dreserva_df)
        errors = detect_errors(reserva_df, rentabilidad_df)
        for err in errors:
            err["Compañia"] = company

        all_serv_rows.extend(serv_rows)
        all_errors.extend(errors)
        total_serv_count += serv_count

        print(f"  {label}: SERV={serv_count}")

    # ── Generar Weekly ───────────────────────────────────────────────────
    output_dir = os.path.join(BASE_DIR, "output")
    os.makedirs(output_dir, exist_ok=True)
    output_xlsx = os.path.join(output_dir, f"Week_{week_num}_{today.year}_VERIFY.xlsx")

    if not os.path.exists(TEMPLATE_PATH):
        print(f"  ERROR: Template no encontrado: {TEMPLATE_PATH}")
        sys.exit(1)

    print(f"\n  Generando Weekly...")
    print(f"  DATA: {len(all_data_rows)} filas (Book 1.xlsx)")
    print(f"  SERV: {total_serv_count} filas (CSVs)")

    wb = generate_weekly_excel(
        TEMPLATE_PATH, output_xlsx, all_data_rows, all_serv_rows,
        total_serv_count, TEMPLATE_FX
    )

    # Booking Window (acumulativo: busca output de semana anterior)
    prev_verify = os.path.join(output_dir, f"Week_{week_num - 1}_{today.year}_VERIFY.xlsx")
    combined_matrix = build_booking_matrix(all_data_rows)
    write_booking_to_excel(wb, combined_matrix, week_num=week_num, prev_output=prev_verify)

    # Diferencia %
    n_weeks = write_diferencia_pct(wb, all_data_rows)
    print(f"  Diferencia %: {n_weeks} semanas")

    # Errores
    n_errors = write_errores_sheet(wb, all_errors)
    if n_errors:
        print(f"  Errores: {n_errors}")

    # ── Filtros de pivots ────────────────────────────────────────────────
    print(f"\n  Actualizando filtros de pivots...")
    n_updated = update_pivot_week_filters(wb, week_num)
    print(f"  {n_updated} pivots actualizados (Weekly SL y LLC → semanas 1-{week_num})")

    wb.save(output_xlsx)
    wb.close()

    print(f"\n  Archivo: {output_xlsx}")
    print(f"  DATA de: {BOOK1_PATH}")
    print(f"═══ Verificación completada ═══")


if __name__ == "__main__":
    main()
