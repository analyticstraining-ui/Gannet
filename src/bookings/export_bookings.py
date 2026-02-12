"""
Exporta los datos de reservas (hoja DATA) como Excel independiente por entidad.
"""

from datetime import datetime

from openpyxl import Workbook


# Headers matching the DATA sheet in the Weekly Report
HEADERS = [
    "Compania", "folio", "cerrada", "fecha", "fecha_inicio", "fecha_fin",
    "vendedor", "Semana", "usuarios_invitados", "total_cliente", "moneda",
    "Total Venta EUR", "Total Venta USD", "Rentabilidad",
    "Rentabilidad en EUR", "Rentabilidad en USD", "% Rentabilidad",
    "Mes", "Ano", "Mes Inicio", "Ano Inicio",
    "Fecha 45 dias fin", "mes 45 dias fin", "Ano 45 dias fin",
    "",  # Y (empty)
    "Observaciones",
]

COL_ORDER = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K",
    "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
    "V", "W", "X", None, "Z",
]


def export_bookings_xlsx(data_rows, output_path, company):
    """Export DATA rows for a single entity to a standalone Excel file.

    Args:
        data_rows: List of row dicts (same format as DATA sheet).
        output_path: Path for the output .xlsx file.
        company: "SL" or "LLC" (for the filename/title).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Bookings {company}"

    # Header row
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    # Data rows
    for row_data in data_rows:
        row = []
        for col_key in COL_ORDER:
            if col_key is None:
                row.append("")
            else:
                row.append(row_data.get(col_key))
        ws.append(row)

    # Format columns
    for row in range(2, len(data_rows) + 2):
        # Dates: D(4), E(5), F(6), V(22)
        for col in [4, 5, 6, 22]:
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, datetime):
                cell.number_format = "MM-DD-YY"
        # Numbers: J(10), L(12), M(13), N(14), O(15), P(16)
        for col in [10, 12, 13, 14, 15, 16]:
            cell = ws.cell(row, col)
            if cell.value is not None:
                cell.number_format = '0'
        # Percentage: Q(17)
        cell = ws.cell(row, 17)
        if cell.value is not None:
            cell.number_format = '0.00%'

    wb.save(output_path)
    wb.close()
    print(f"  Bookings {company} exportado: {output_path}")
