"""
Módulo Booking Window: genera la matriz semana x mes y exporta a Excel.

Lógica acumulativa: cada semana copia los datos de la semana anterior
y añade los datos nuevos de la semana actual.
"""

import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from config import MONTH_NAMES_ES

# USD format matching the template
USD_FMT = '[$$-409]#,##0'

# Columnas de datos: C-N (2026), O (total 2026), P-AA (2027), AB (total 2027)
_MONTH_COLS_2026 = range(3, 15)   # C-N
_TOTAL_COL_2026 = 15              # O
_MONTH_COLS_2027 = range(16, 28)  # P-AA
_TOTAL_COL_2027 = 28              # AB
_ALL_DATA_COLS = range(3, 29)     # C-AB


def _week_to_row(week_num):
    """Convert week number to row in Booking Window sheet.
    Week 53 is row 2, Week 52 is row 4, ..., Week 1 is row 106."""
    return 2 + (53 - week_num) * 2


def _carry_forward_booking(ws_bw, prev_path):
    """Copia datos del Booking Window de la semana anterior al sheet actual.

    Solo copia semanas que no existen ya en el sheet actual (del template).
    """
    if not os.path.isfile(prev_path):
        return 0

    prev_wb = load_workbook(prev_path, data_only=True)
    if "Booking Window 2026" not in prev_wb.sheetnames:
        prev_wb.close()
        return 0

    prev_ws = prev_wb["Booking Window 2026"]
    carried = 0

    for week in range(1, 54):
        value_row = _week_to_row(week)
        pct_row = value_row + 1

        # Saltar si el sheet actual ya tiene datos (del template)
        has_current = any(
            ws_bw.cell(value_row, col).value is not None
            for col in _MONTH_COLS_2026
        )
        if has_current:
            continue

        # Verificar si la semana anterior tiene datos
        has_prev = any(
            prev_ws.cell(value_row, col).value is not None
            for col in _MONTH_COLS_2026
        )
        if not has_prev:
            continue

        # Copiar valores y porcentajes
        for col in _ALL_DATA_COLS:
            val = prev_ws.cell(value_row, col).value
            if val is not None:
                cell = ws_bw.cell(value_row, col)
                cell.value = val
                cell.number_format = USD_FMT

            pct_val = prev_ws.cell(pct_row, col).value
            if pct_val is not None:
                cell = ws_bw.cell(pct_row, col)
                cell.value = pct_val
                cell.number_format = '0%'

        carried += 1

    prev_wb.close()
    return carried


def build_booking_matrix(data_rows):
    """Build the booking window matrix: for each sale week, sum Total Venta USD
    by departure month.

    Args:
        data_rows: List of DATA sheet row dicts.

    Returns:
        Dict {week: {month_key: total_usd}} where month_key 1-12 = 2026,
        14-25 = 2027 months.
    """
    matrix = {}

    for r in data_rows:
        fecha_inicio = r["E"]  # departure date
        total_usd = r["M"]    # Total Venta USD
        fecha = r["D"]         # sale date

        if not fecha or not fecha_inicio:
            continue

        sale_week = int(fecha.strftime('%W')) + 1
        dep_year = fecha_inicio.year
        dep_month = fecha_inicio.month

        if sale_week not in matrix:
            matrix[sale_week] = {}

        if dep_year == 2026:
            matrix[sale_week][dep_month] = (
                matrix[sale_week].get(dep_month, 0) + total_usd
            )
        elif dep_year == 2027:
            matrix[sale_week][dep_month + 13] = (
                matrix[sale_week].get(dep_month + 13, 0) + total_usd
            )

    return matrix


def write_booking_to_excel(wb, matrix, week_num=None, prev_output=None):
    """Write the booking window matrix into the 'Booking Window 2026' sheet.

    Lógica acumulativa:
    1. Copiar datos de la semana anterior (prev_output) si existe
    2. Escribir solo semanas nuevas desde la matrix
    3. Ocultar filas futuras (> week_num)

    Args:
        wb: Open openpyxl Workbook.
        matrix: Booking matrix from build_booking_matrix().
        week_num: Semana actual (para ocultar filas futuras).
        prev_output: Ruta al output de la semana anterior (para carry-forward).
    """
    ws_bw = wb["Booking Window 2026"]

    # Paso 1: Carry-forward desde la semana anterior
    if prev_output:
        n_carried = _carry_forward_booking(ws_bw, prev_output)
        if n_carried:
            print(f"  Booking Window: {n_carried} semanas copiadas de {os.path.basename(prev_output)}")

    # Paso 2: Escribir semanas nuevas
    weeks_with_data = set()

    for week in range(1, 54):
        value_row = _week_to_row(week)

        # Verificar si ya tiene datos (del template o carry-forward)
        has_existing = any(
            ws_bw.cell(value_row, col).value is not None
            for col in _MONTH_COLS_2026
        )

        if has_existing:
            weeks_with_data.add(week)
            continue

        # No escribir semanas futuras
        if week_num and week > week_num:
            continue

        # Escribir datos nuevos si existen en la matrix
        if week not in matrix:
            continue

        month_totals = matrix[week]
        total_2026 = 0
        total_2027 = 0

        # Write 2026 months (cols C-N = 3-14)
        for month in range(1, 13):
            val = month_totals.get(month, 0)
            if val:
                col_idx = month + 2
                cell = ws_bw.cell(value_row, col_idx)
                cell.value = round(val, 2)
                cell.number_format = USD_FMT
                total_2026 += val

        # Write total 2026 (col O = 15)
        if total_2026:
            cell = ws_bw.cell(value_row, 15)
            cell.value = round(total_2026, 2)
            cell.number_format = USD_FMT

        # Write 2027 months (cols P-AA = 16-27)
        for month_key in range(14, 26):
            val = month_totals.get(month_key, 0)
            if val:
                month_2027 = month_key - 13
                col_idx = month_2027 + 15
                cell = ws_bw.cell(value_row, col_idx)
                cell.value = round(val, 2)
                cell.number_format = USD_FMT
                total_2027 += val

        # Write total 2027 (col AB = 28)
        if total_2027:
            cell = ws_bw.cell(value_row, 28)
            cell.value = round(total_2027, 2)
            cell.number_format = USD_FMT

        # Write percentage formulas in the row below
        pct_row = value_row + 1
        if total_2026:
            for col in range(3, 15):
                cell = ws_bw.cell(pct_row, col)
                cell.value = (
                    f"={get_column_letter(col)}{value_row}/$O${value_row}"
                )
                cell.number_format = '0%'

        weeks_with_data.add(week)

    # Paso 3: Gestionar visibilidad de filas
    if week_num:
        shown = []
        for week in range(1, 54):
            value_row = _week_to_row(week)
            pct_row = value_row + 1
            visible = week <= week_num and week in weeks_with_data
            ws_bw.row_dimensions[value_row].hidden = not visible
            ws_bw.row_dimensions[pct_row].hidden = not visible
            if visible:
                shown.append(week)
        if shown:
            print(f"  Booking Window: semanas {min(shown)}-{max(shown)} visibles, resto oculto")


def export_booking_xlsx(matrix, output_path):
    """Export the booking window matrix as a standalone Excel file.

    Args:
        matrix: Booking matrix from build_booking_matrix().
        output_path: Path for the output .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Booking Window"

    # Header
    months_2026 = [MONTH_NAMES_ES[m].upper()[:3] + " 2026" for m in range(1, 13)]
    months_2027 = [MONTH_NAMES_ES[m].upper()[:3] + " 2027" for m in range(1, 13)]
    header = ["SEMANA"] + months_2026 + ["TOTAL 2026"] + months_2027 + ["TOTAL 2027"]
    ws.append(header)

    # Bold header
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    # Data rows (sorted by week descending, like the template)
    for week_num in sorted(matrix.keys(), reverse=True):
        month_totals = matrix[week_num]
        row = [f"WEEK {week_num}"]

        total_2026 = 0
        for month in range(1, 13):
            val = round(month_totals.get(month, 0), 2)
            row.append(val)
            total_2026 += val
        row.append(round(total_2026, 2))

        total_2027 = 0
        for month_key in range(14, 26):
            val = round(month_totals.get(month_key, 0), 2)
            row.append(val)
            total_2027 += val
        row.append(round(total_2027, 2))

        ws.append(row)

    # Format USD columns (B onwards)
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.number_format = USD_FMT

    wb.save(output_path)
    wb.close()
    print(f"  Booking Window Excel exportado: {output_path}")
