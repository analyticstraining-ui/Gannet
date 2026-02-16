"""
Genera reportes mensuales individuales por Travel Advisor.

Cada TA con reservas en el mes recibe su propio Excel con:
  - Hoja "Reporte": dos tablas (venta por mes de salida y por mes de calendario)
  - Hoja "DATA": todas las reservas del TA
"""

import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

MONTH_NAMES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

USD_FMT = '[$$-409]#,##0'
PCT_FMT = '0.0%'
DATE_FMT = 'DD/MM/YY'

BOLD = Font(name="Arial", bold=True, size=10)
NORMAL = Font(name="Arial", size=10)
HDR_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Columnas de la hoja DATA individual
DATA_COLS = [
    ("A", "Compañia"), ("B", "folio"), ("C", "cerrada"),
    ("D", "fecha"), ("E", "fecha_inicio"), ("F", "fecha_fin"),
    ("G", "vendedor"), ("H", "Semana"), ("I", "usuarios_invitados"),
    ("J", "total_cliente"), ("K", "moneda"),
    ("L", "Total Venta EUR"), ("M", "Total Venta USD"),
    ("N", "Rentabilidad"), ("O", "Rentabilidad en EUR"),
    ("P", "Rentabilidad en USD"), ("Q", "% Rentabilidad"),
    ("R", "Mes"), ("S", "Año"), ("T", "Mes Inicio"), ("U", "Año Inicio"),
]


def generate_individual_reports(data_rows, output_base_dir, year=2026, month=1):
    """
    Genera un reporte Excel individual por cada TA con reservas en el mes/año dado.

    Args:
        data_rows: Lista de dicts con claves A-Z (salida de build_data_rows).
        output_base_dir: Directorio base para crear la carpeta de reportes.
        year: Año a filtrar.
        month: Mes a filtrar.

    Returns:
        Número de reportes generados.
    """
    month_name = MONTH_NAMES_ES.get(month, str(month))
    folder_name = f"Reportes_Individuales_{month_name}_{year}"
    base_dir = os.path.join(output_base_dir, folder_name)

    # Identificar TAs con reservas en el mes/año (por fecha de calendario)
    tas_in_month = set()
    for r in data_rows:
        if r.get("R") == month and r.get("S") == year:
            vend = r.get("G")
            if vend:
                tas_in_month.add(vend)

    if not tas_in_month:
        print(f"  No hay TAs con reservas en {month_name} {year}")
        return 0

    print(f"  {len(tas_in_month)} TAs con reservas en {month_name} {year}")

    count = 0
    for ta in sorted(tas_in_month):
        # Todas las reservas del TA (para hoja DATA)
        ta_all = [r for r in data_rows if r.get("G") == ta]

        # Reservas del TA en el año (para tablas del reporte)
        ta_year = [r for r in ta_all if r.get("S") == year]

        # Crear carpeta del TA
        ta_dir = os.path.join(base_dir, ta)
        os.makedirs(ta_dir, exist_ok=True)
        filepath = os.path.join(ta_dir, f"{ta}_{month_name}_{year}.xlsx")

        wb = Workbook()
        _write_reporte_sheet(wb, ta, ta_year, year)
        _write_data_sheet(wb, ta_all)
        wb.save(filepath)
        wb.close()
        count += 1

    print(f"  {count} reportes generados en {base_dir}")
    return count


def _write_reporte_sheet(wb, ta_name, ta_year_rows, year):
    """Escribe la hoja Reporte con las dos tablas."""
    ws = wb.active
    ws.title = "Reporte"

    # ── Tabla 1: Venta por mes de salida ──────────────────────────
    row = 2
    ws.cell(row, 1, "Venta por mes de salida de viajes").font = BOLD

    row = 3
    ws.cell(row, 1, "Suma de Total Venta USD").font = BOLD
    ws.cell(row, 2, str(year)).font = BOLD

    # Agrupar por Mes Inicio donde Año Inicio == year
    salida = defaultdict(float)
    for r in ta_year_rows:
        if r.get("U") == year and r.get("T") is not None:
            mes_inicio = int(r["T"])
            salida[mes_inicio] += float(r.get("M") or 0)

    meses_salida = sorted(salida.keys())

    # Headers
    row = 4
    ws.cell(row, 1, "Etiquetas de fila").font = BOLD
    ws.cell(row, 1).fill = HDR_FILL
    for ci, mes in enumerate(meses_salida, 2):
        ws.cell(row, ci, mes).font = BOLD
        ws.cell(row, ci).fill = HDR_FILL
    col_total = len(meses_salida) + 2
    col_grand = col_total + 1
    ws.cell(row, col_total, f"{year} Total").font = BOLD
    ws.cell(row, col_total).fill = HDR_FILL
    ws.cell(row, col_grand, "Grand Total").font = BOLD
    ws.cell(row, col_grand).fill = HDR_FILL

    # Datos del TA
    row = 5
    ws.cell(row, 1, ta_name).font = NORMAL
    total_salida = 0
    for ci, mes in enumerate(meses_salida, 2):
        val = salida[mes]
        total_salida += val
        ws.cell(row, ci, val).number_format = USD_FMT
    ws.cell(row, col_total, total_salida).number_format = USD_FMT
    ws.cell(row, col_grand, total_salida).number_format = USD_FMT
    ws.cell(row, col_grand).font = BOLD

    # Total general
    row = 6
    ws.cell(row, 1, "Total general").font = BOLD
    for ci, mes in enumerate(meses_salida, 2):
        c = ws.cell(row, ci, salida[mes])
        c.number_format = USD_FMT
        c.font = BOLD
    ws.cell(row, col_total, total_salida).number_format = USD_FMT
    ws.cell(row, col_total).font = BOLD

    # ── Tabla 2: Venta por mes de calendario ──────────────────────
    row = 10
    ws.cell(row, 1, "Venta por mes de calendario").font = BOLD

    row = 11
    ws.cell(row, 1, "Suma de Total Venta USD").font = BOLD
    ws.cell(row, 2, str(year)).font = BOLD

    # Agrupar por Mes donde Año == year
    cal_venta = defaultdict(float)
    cal_rent = defaultdict(float)
    for r in ta_year_rows:
        if r.get("S") == year and r.get("R") is not None:
            mes = int(r["R"])
            cal_venta[mes] += float(r.get("M") or 0)
            cal_rent[mes] += float(r.get("P") or 0)

    meses_cal = sorted(cal_venta.keys())
    total_venta_cal = sum(cal_venta.values())
    total_rent_cal = sum(cal_rent.values())
    rent_pct = total_rent_cal / total_venta_cal if total_venta_cal > 0 else 0

    # Headers
    row = 12
    ws.cell(row, 1, "Etiquetas de fila").font = BOLD
    ws.cell(row, 1).fill = HDR_FILL
    for ci, mes in enumerate(meses_cal, 2):
        ws.cell(row, ci, mes).font = BOLD
        ws.cell(row, ci).fill = HDR_FILL
    col_t2_total = len(meses_cal) + 2
    col_t2_grand = col_t2_total + 1
    col_t2_benef = col_t2_grand + 1
    col_t2_rent = col_t2_benef + 1
    for ci, label in [(col_t2_total, f"{year} Total"),
                       (col_t2_grand, "Grand Total"),
                       (col_t2_benef, "Beneficio USD"),
                       (col_t2_rent, "Rentabilidad %")]:
        ws.cell(row, ci, label).font = BOLD
        ws.cell(row, ci).fill = HDR_FILL

    # Datos del TA
    row = 13
    ws.cell(row, 1, ta_name).font = NORMAL
    for ci, mes in enumerate(meses_cal, 2):
        ws.cell(row, ci, cal_venta[mes]).number_format = USD_FMT
    ws.cell(row, col_t2_total, total_venta_cal).number_format = USD_FMT
    ws.cell(row, col_t2_total).font = BOLD
    ws.cell(row, col_t2_grand, total_venta_cal).number_format = USD_FMT
    ws.cell(row, col_t2_grand).font = BOLD
    ws.cell(row, col_t2_benef, total_rent_cal).number_format = USD_FMT
    ws.cell(row, col_t2_rent, rent_pct).number_format = PCT_FMT

    # Total general
    row = 14
    ws.cell(row, 1, "Total general").font = BOLD
    for ci, mes in enumerate(meses_cal, 2):
        c = ws.cell(row, ci, cal_venta[mes])
        c.number_format = USD_FMT
        c.font = BOLD
    ws.cell(row, col_t2_total, total_venta_cal).number_format = USD_FMT
    ws.cell(row, col_t2_total).font = BOLD

    # Ajustar anchos
    ws.column_dimensions["A"].width = 22
    for col in range(2, col_grand + 5):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _write_data_sheet(wb, ta_all_rows):
    """Escribe la hoja DATA con todas las reservas del TA."""
    ws = wb.create_sheet("DATA")

    # Headers
    for ci, (_, header) in enumerate(DATA_COLS, 1):
        cell = ws.cell(1, ci, header)
        cell.font = BOLD
        cell.fill = HDR_FILL

    # Datos
    for ri, row_data in enumerate(ta_all_rows, 2):
        for ci, (col_letter, _) in enumerate(DATA_COLS, 1):
            val = row_data.get(col_letter)
            if val is not None:
                cell = ws.cell(ri, ci, val)
                # Formato por tipo de columna
                if col_letter in ("D", "E", "F"):
                    cell.number_format = DATE_FMT
                elif col_letter in ("J", "L", "M", "N", "O", "P"):
                    cell.number_format = '#,##0.00'
                elif col_letter == "Q":
                    cell.number_format = PCT_FMT

    # Ajustar anchos
    widths = {"A": 10, "B": 8, "C": 7, "D": 12, "E": 12, "F": 12,
              "G": 14, "H": 8, "I": 10, "J": 14, "K": 8,
              "L": 16, "M": 16, "N": 14, "O": 16, "P": 16,
              "Q": 14, "R": 6, "S": 6, "T": 10, "U": 10}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w
