"""
Dashboard de Análisis — genera Dashboard_Insights.xlsx
con 10 hojas, cada una con tabla resumen + gráfico.
"""

from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


# ── Estilos ──────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Arial", bold=True, size=14)
HDR_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
DATA_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
USD = '#,##0'
PCT = '0.0%'
INT = '#,##0'
CW = 24   # chart width (cm)
CH = 15   # chart height (cm)


# ── Helpers ──────────────────────────────────────────────────────────

def _sheet(wb, name, title, headers, widths=None):
    """Create sheet with title + headers. Returns (ws, first_data_row=4)."""
    ws = wb.create_sheet(name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    return ws, 4


def _rows(ws, sr, data, fmts=None):
    """Write rows with alternating fills. Returns row after last data row."""
    for i, rd in enumerate(data):
        r = sr + i
        for c, val in enumerate(rd, 1):
            cell = ws.cell(r, c, val)
            cell.font = DATA_FONT
            cell.border = BORDER
            if fmts and c <= len(fmts) and fmts[c - 1]:
                cell.number_format = fmts[c - 1]
            if i % 2 == 1:
                cell.fill = ALT_FILL
    return sr + len(data)


def _fix_axes(chart, rotate_x=0):
    """Ensure both axes are visible with proper labels."""
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    if rotate_x:
        chart.x_axis.tickLblRot = rotate_x


def _fix_combo(primary, secondary, rotate_x=0):
    """Fix axes for combined bar+line charts."""
    primary.x_axis.delete = False
    primary.y_axis.delete = False
    primary.x_axis.tickLblPos = "low"
    if rotate_x:
        primary.x_axis.tickLblRot = rotate_x
    secondary.y_axis.axId = 200
    secondary.y_axis.crosses = "min"
    secondary.y_axis.delete = False
    primary += secondary


# ── Sheet 1: Rendimiento Vendedor ───────────────────────────────────

def _s1(wb, rows):
    hdrs = ["Vendedor", "Total Venta USD", "Rentabilidad USD",
            "% Rent. Prom.", "Nº Cotizaciones", "Ticket Prom. USD"]
    ws, dr = _sheet(wb, "Rendimiento Vendedor",
                    "Rendimiento por Vendedor", hdrs,
                    [25, 18, 18, 15, 16, 18])

    st = defaultdict(lambda: {"usd": 0, "rent": 0, "pcts": [], "n": 0})
    for r in rows:
        v = r["G"] or "Sin vendedor"
        st[v]["usd"] += r["M"] or 0
        st[v]["rent"] += r["P"] or 0
        if r["Q"] is not None:
            st[v]["pcts"].append(r["Q"])
        st[v]["n"] += 1

    top = sorted(st.items(), key=lambda x: -x[1]["usd"])[:15]
    tbl = []
    for v, s in top:
        ap = sum(s["pcts"]) / len(s["pcts"]) if s["pcts"] else 0
        tk = s["usd"] / s["n"] if s["n"] else 0
        tbl.append((v, s["usd"], s["rent"], ap, s["n"], tk))

    end = _rows(ws, dr, tbl, [None, USD, USD, PCT, INT, USD])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    vals = Reference(ws, min_col=2, min_row=3, max_row=end - 1)
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Top Vendedores — Venta USD"
    chart.y_axis.numFmt = USD
    chart.width = CW
    chart.height = CH
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    _fix_axes(chart)
    # For horizontal bars, the "x_axis" is actually the value axis
    chart.x_axis.numFmt = USD
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.tickLblPos = "low"
    ws.add_chart(chart, f"A{end + 2}")


# ── Sheet 2: Evolución Semanal ──────────────────────────────────────
def _s2(wb, rows):
    hdrs = ["Semana", "Venta USD 2025", "Venta USD 2026",
            "Variación %", "Rent USD 2025", "Rent USD 2026"]
    ws, dr = _sheet(wb, "Evolución Semanal",
                    "Evolución Semanal 2026 vs 2025", hdrs,
                    [12, 18, 18, 14, 18, 18])

    bwy = defaultdict(lambda: defaultdict(lambda: {"usd": 0, "rent": 0}))
    for r in rows:
        w, y = r["H"], r["S"]
        if w and y in (2025, 2026):
            bwy[w][y]["usd"] += r["M"] or 0
            bwy[w][y]["rent"] += r["P"] or 0

    weeks = sorted(w for w in bwy if 2026 in bwy[w])
    tbl = []
    for w in weeks:
        v25 = bwy[w][2025]["usd"]
        v26 = bwy[w][2026]["usd"]
        var = (v26 - v25) / v25 if v25 else 0
        tbl.append((f"S{w}", v25, v26, var, bwy[w][2025]["rent"], bwy[w][2026]["rent"]))

    if not tbl:
        return

    end = _rows(ws, dr, tbl, [None, USD, USD, PCT, USD, USD])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    v25 = Reference(ws, min_col=2, min_row=3, max_row=end - 1)
    v26 = Reference(ws, min_col=3, min_row=3, max_row=end - 1)

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Venta Semanal 2025 vs 2026 (USD)"
    bar.width = CW
    bar.height = CH

    # --- FIX: Left Y-axis for USD bars ---
    bar.y_axis.title = "USD"
    bar.y_axis.numFmt = USD
    bar.y_axis.delete = False
    bar.y_axis.majorGridlines = None  # Remove gridlines to reduce clutter
    bar.y_axis.axId = 100

    bar.add_data(v25, titles_from_data=True)
    bar.add_data(v26, titles_from_data=True)
    bar.set_categories(cats)

    # --- FIX: Right Y-axis for Variación % line ---
    line = LineChart()
    vv = Reference(ws, min_col=4, min_row=3, max_row=end - 1)
    line.add_data(vv, titles_from_data=True)

    line.y_axis.title = "Variación %"
    line.y_axis.numFmt = PCT
    line.y_axis.delete = False
    line.y_axis.crosses = "max"   # Push to the RIGHT side
    line.y_axis.axId = 200

    # Combine charts
    bar += line

    # Style the line to make it distinct
    s = bar.series[-1]
    s.graphicalProperties.line.width = 25000

    ws.add_chart(bar, f"A{end + 2}")


# ── Sheet 3: Distribución Rentabilidad ──────────────────────────────
def _s3(wb, rows):
    hdrs = ["Rango", "Nº Cotizaciones", "% del Total", "Venta USD Promedio"]
    ws, dr = _sheet(wb, "Distribución Rentabilidad",
                    "Distribución de Rentabilidad", hdrs,
                    [18, 16, 14, 18])

    n_total = len(rows)
    ranges = [
        ("<0%", lambda q: q < 0),
        ("0%", lambda q: q == 0),
        ("0.1-5%", lambda q: 0 < q <= 0.05),
        ("5-10%", lambda q: 0.05 < q <= 0.10),
        ("10-15%", lambda q: 0.10 < q <= 0.15),
        ("15-20%", lambda q: 0.15 < q <= 0.20),
        (">20%", lambda q: q > 0.20),
    ]

    tbl = []
    for label, cond in ranges:
        m = [r for r in rows if r["Q"] is not None and cond(r["Q"])]
        n = len(m)
        pct = n / n_total if n_total else 0
        avg = sum(r["M"] or 0 for r in m) / n if n else 0
        tbl.append((label, n, pct, avg))

    end = _rows(ws, dr, tbl, [None, INT, PCT, USD])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    vc = Reference(ws, min_col=2, min_row=3, max_row=end - 1)

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Distribución de Rentabilidad"
    bar.width = CW
    bar.height = CH

    # --- FIX: Configure the LEFT Y-axis (bar chart) clearly ---
    bar.y_axis.title = "Nº Cotizaciones"
    bar.y_axis.numFmt = INT
    bar.y_axis.delete = False
    # Prevent axis labels from overlapping by removing gridline clutter
    bar.y_axis.majorGridlines = None

    bar.add_data(vc, titles_from_data=True)
    bar.set_categories(cats)

    # --- FIX: Configure the line chart for the RIGHT Y-axis ---
    line = LineChart()
    va = Reference(ws, min_col=4, min_row=3, max_row=end - 1)
    line.add_data(va, titles_from_data=True)

    line.y_axis.title = "Venta USD Promedio"
    line.y_axis.numFmt = USD
    line.y_axis.delete = False
    # Cross the right axis at the max so it stays on the right side
    line.y_axis.crosses = "max"

    # --- FIX: Ensure axes don't overlap by using axId separation ---
    # Give each chart axis a unique ID so Excel renders them independently
    bar.y_axis.axId = 100
    line.y_axis.axId = 200

    # Combine: add line series to bar chart, map them to the secondary axis
    bar += line

    # Style the line series so it's clearly distinct
    s = bar.series[-1]
    s.graphicalProperties.line.width = 25000  # thicker line

    ws.add_chart(bar, f"A{end + 2}")


# ── Sheet 4: Booking Window ─────────────────────────────────────────
def _s4(wb, rows):
    hdrs = ["Anticipación", "Nº Cotizaciones", "Venta USD Total",
            "Venta USD Promedio", "Rent % Promedio"]
    ws, dr = _sheet(wb, "Booking Window",
                    "Booking Window — Anticipación de Reserva", hdrs,
                    [18, 16, 18, 18, 15])

    buckets = [
        ("0-30d", 0, 31),
        ("31-60d", 31, 61),
        ("61-90d", 61, 91),
        ("91-180d", 91, 181),
        ("181-365d", 181, 366),
        (">365d", 366, 99999),
    ]

    bd = defaultdict(lambda: {"n": 0, "usd": 0, "pcts": []})
    for r in rows:
        if r["E"] and r["D"]:
            try:
                diff = (r["E"] - r["D"]).days
                for label, lo, hi in buckets:
                    if lo <= diff < hi:
                        bd[label]["n"] += 1
                        bd[label]["usd"] += r["M"] or 0
                        if r["Q"] is not None:
                            bd[label]["pcts"].append(r["Q"])
                        break
            except (TypeError, AttributeError):
                pass

    tbl = []
    for label, _, _ in buckets:
        s = bd[label]
        av = s["usd"] / s["n"] if s["n"] else 0
        ap = sum(s["pcts"]) / len(s["pcts"]) if s["pcts"] else 0
        tbl.append((label, s["n"], s["usd"], av, ap))

    end = _rows(ws, dr, tbl, [None, INT, USD, USD, PCT])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    vu = Reference(ws, min_col=3, min_row=3, max_row=end - 1)

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Booking Window"
    bar.width = CW
    bar.height = CH

    # Left Y-axis: Venta USD
    bar.y_axis.title = "Venta USD"
    bar.y_axis.numFmt = USD
    bar.y_axis.delete = False
    bar.y_axis.majorGridlines = None
    bar.y_axis.axId = 100

    bar.add_data(vu, titles_from_data=True)
    bar.set_categories(cats)

    # Right Y-axis: Rent %
    line = LineChart()
    vp = Reference(ws, min_col=5, min_row=3, max_row=end - 1)
    line.add_data(vp, titles_from_data=True)

    line.y_axis.title = "Rent %"
    line.y_axis.numFmt = PCT
    line.y_axis.delete = False
    line.y_axis.crosses = "max"
    line.y_axis.axId = 200

    # Combine
    bar += line

    s = bar.series[-1]
    s.graphicalProperties.line.width = 25000

    ws.add_chart(bar, f"A{end + 2}")

# ── Sheet 5: Estacionalidad ─────────────────────────────────────────

def _s5(wb, rows):
    hdrs = ["Mes", "Total Venta USD", "Nº Cotizaciones",
            "Ticket Promedio", "Rent % Promedio"]
    ws, dr = _sheet(wb, "Estacionalidad Mes Salida",
                    "Estacionalidad — Mes de Salida", hdrs,
                    [14, 18, 16, 18, 15])

    months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
              "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    filtered = [r for r in rows if r.get("U") in (2025, 2026)]
    bm = defaultdict(lambda: {"usd": 0, "n": 0, "pcts": []})
    for r in filtered:
        m = r.get("T")
        if m and 1 <= m <= 12:
            bm[m]["usd"] += r["M"] or 0
            bm[m]["n"] += 1
            if r["Q"] is not None:
                bm[m]["pcts"].append(r["Q"])

    tbl = []
    for m in range(1, 13):
        s = bm[m]
        tk = s["usd"] / s["n"] if s["n"] else 0
        ap = sum(s["pcts"]) / len(s["pcts"]) if s["pcts"] else 0
        tbl.append((months[m - 1], s["usd"], s["n"], tk, ap))

    end = _rows(ws, dr, tbl, [None, USD, INT, USD, PCT])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    vu = Reference(ws, min_col=2, min_row=3, max_row=end - 1)

    line = LineChart()
    line.title = "Estacionalidad — Mes de Salida"
    line.style = 10
    line.y_axis.title = "Venta USD"
    line.y_axis.numFmt = USD
    line.width = CW
    line.height = CH
    line.add_data(vu, titles_from_data=True)
    line.set_categories(cats)

    bar = BarChart()
    vc = Reference(ws, min_col=3, min_row=3, max_row=end - 1)
    bar.add_data(vc, titles_from_data=True)
    bar.y_axis.title = "Nº Cotizaciones"
    bar.y_axis.numFmt = INT

    _fix_combo(line, bar)
    ws.add_chart(line, f"A{end + 2}")


# ── Sheet 6: Mix Monedas ────────────────────────────────────────────

def _s6(wb, rows):
    hdrs = ["Moneda", "Nº Cotizaciones", "Total Venta USD", "% del Total"]
    ws, dr = _sheet(wb, "Mix Monedas",
                    "Mix de Monedas", hdrs,
                    [14, 16, 18, 14])

    total_usd = sum(r["M"] or 0 for r in rows)
    bc = defaultdict(lambda: {"n": 0, "usd": 0})
    for r in rows:
        cur = r["K"] or "N/A"
        bc[cur]["n"] += 1
        bc[cur]["usd"] += r["M"] or 0

    tbl = []
    for cur, s in sorted(bc.items(), key=lambda x: -x[1]["usd"]):
        pct = s["usd"] / total_usd if total_usd else 0
        tbl.append((cur, s["n"], s["usd"], pct))

    end = _rows(ws, dr, tbl, [None, INT, USD, PCT])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    vals = Reference(ws, min_col=3, min_row=3, max_row=end - 1)

    pie = PieChart()
    pie.title = "Mix de Monedas"
    pie.style = 10
    pie.width = 20
    pie.height = 16
    pie.add_data(vals, titles_from_data=True)
    pie.set_categories(cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = False
    pie.dataLabels.showVal = False
    pie.legend.position = "b"

    ws.add_chart(pie, f"A{end + 2}")


# ── Sheet 7: Rentabilidad por Servicio ──────────────────────────────

def _s7(wb, serv_rows, fx):
    hdrs = ["Tipo Servicio", "Subtotal USD", "Comisión USD",
            "% Rentabilidad", "Nº Servicios"]
    ws, dr = _sheet(wb, "Rentabilidad por Servicio",
                    "Rentabilidad por Tipo de Servicio", hdrs,
                    [18, 18, 18, 15, 14])

    main_types = {"HTL", "VLO", "PKT", "TRF", "HPS", "ATR"}
    bt = defaultdict(lambda: {"sub": 0, "com": 0, "n": 0})

    for r in serv_rows:
        tipo = str(r.get("I") or "OTR").strip()
        if tipo not in main_types:
            tipo = "Otros"
        moneda = str(r.get("K") or "EUR").strip()
        fx_usd = fx.get(moneda, fx.get("EUR", {"USD": 1.0})).get("USD", 1.0)
        bt[tipo]["sub"] += float(r.get("L") or 0) * fx_usd
        bt[tipo]["com"] += float(r.get("O") or 0) * fx_usd
        bt[tipo]["n"] += 1

    tbl = []
    for tipo, s in sorted(bt.items(), key=lambda x: -x[1]["sub"]):
        pct = s["com"] / s["sub"] if s["sub"] else 0
        tbl.append((tipo, round(s["sub"]), round(s["com"]), pct, s["n"]))

    end = _rows(ws, dr, tbl, [None, USD, USD, PCT, INT])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    vs = Reference(ws, min_col=2, min_row=3, max_row=end - 1)

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Rentabilidad por Tipo de Servicio"
    bar.y_axis.title = "Subtotal USD"
    bar.y_axis.numFmt = USD
    bar.width = CW
    bar.height = CH
    bar.add_data(vs, titles_from_data=True)
    bar.set_categories(cats)

    line = LineChart()
    vp = Reference(ws, min_col=4, min_row=3, max_row=end - 1)
    line.add_data(vp, titles_from_data=True)
    line.y_axis.title = "% Rentabilidad"
    line.y_axis.numFmt = PCT

    _fix_combo(bar, line)
    ws.add_chart(bar, f"A{end + 2}")


# ── Sheet 8: Concentración Ventas ───────────────────────────────────

def _s8(wb, rows):
    hdrs = ["Folio", "Vendedor", "Venta USD", "% Rentabilidad",
            "Moneda", "Fecha", "% Acumulado"]
    ws, dr = _sheet(wb, "Concentración Ventas",
                    "Concentración de Ventas — Top 20", hdrs,
                    [12, 20, 18, 15, 12, 14, 14])

    total_usd = sum(r["M"] or 0 for r in rows)
    top20 = sorted(rows, key=lambda r: r["M"] or 0, reverse=True)[:20]

    tbl = []
    cumul = 0
    for r in top20:
        cumul += r["M"] or 0
        pa = cumul / total_usd if total_usd else 0
        fs = r["D"].strftime("%Y-%m-%d") if r["D"] else ""
        tbl.append((r["B"], r["G"] or "", r["M"] or 0, r["Q"] or 0,
                    r["K"] or "", fs, pa))

    end = _rows(ws, dr, tbl, [INT, None, USD, PCT, None, None, PCT])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    vu = Reference(ws, min_col=3, min_row=3, max_row=end - 1)

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Concentración de Ventas (Pareto)"
    bar.y_axis.title = "Venta USD"
    bar.y_axis.numFmt = USD
    bar.width = CW
    bar.height = CH
    bar.add_data(vu, titles_from_data=True)
    bar.set_categories(cats)

    line = LineChart()
    va = Reference(ws, min_col=7, min_row=3, max_row=end - 1)
    line.add_data(va, titles_from_data=True)
    line.y_axis.title = "% Acumulado"
    line.y_axis.numFmt = PCT

    _fix_combo(bar, line, rotate_x=-45)
    ws.add_chart(bar, f"A{end + 2}")


# ── Sheet 9: LLC vs SL ──────────────────────────────────────────────

def _s9(wb, rows):
    hdrs = ["Compañía", "Total Venta USD", "Nº Cotizaciones",
            "Ticket Promedio", "Rent % Promedio", "Rent USD Total"]
    ws, dr = _sheet(wb, "LLC vs SL",
                    "LLC vs SL", hdrs,
                    [14, 18, 16, 18, 15, 18])

    bc = defaultdict(lambda: {"usd": 0, "rent": 0, "n": 0, "pcts": []})
    for r in rows:
        c = r["A"]
        bc[c]["usd"] += r["M"] or 0
        bc[c]["rent"] += r["P"] or 0
        bc[c]["n"] += 1
        if r["Q"] is not None:
            bc[c]["pcts"].append(r["Q"])

    tbl = []
    for comp in sorted(bc.keys()):
        s = bc[comp]
        tk = s["usd"] / s["n"] if s["n"] else 0
        ap = sum(s["pcts"]) / len(s["pcts"]) if s["pcts"] else 0
        tbl.append((comp, s["usd"], s["n"], tk, ap, s["rent"]))

    end = _rows(ws, dr, tbl, [None, USD, INT, USD, PCT, USD])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    vu = Reference(ws, min_col=2, min_row=3, max_row=end - 1)
    vr = Reference(ws, min_col=6, min_row=3, max_row=end - 1)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "LLC vs SL — Venta y Rentabilidad (USD)"
    chart.y_axis.title = "USD"
    chart.y_axis.numFmt = USD
    chart.width = CW
    chart.height = CH
    chart.add_data(vu, titles_from_data=True)
    chart.add_data(vr, titles_from_data=True)
    chart.set_categories(cats)
    _fix_axes(chart)

    ws.add_chart(chart, f"A{end + 2}")


# ── Sheet 10: Tasa Cierre ───────────────────────────────────────────

def _s10(wb, rows):
    hdrs = ["Vendedor", "Total", "Cerradas", "Abiertas",
            "Tasa Cierre %", "Venta USD Cerradas", "Venta USD Abiertas"]
    ws, dr = _sheet(wb, "Tasa Cierre",
                    "Tasa de Cierre por Vendedor", hdrs,
                    [25, 10, 12, 12, 14, 18, 18])

    bv = defaultdict(lambda: {"t": 0, "c": 0, "a": 0, "uc": 0, "ua": 0})
    for r in rows:
        v = r["G"] or "Sin vendedor"
        bv[v]["t"] += 1
        if r["C"] == 1:
            bv[v]["c"] += 1
            bv[v]["uc"] += r["M"] or 0
        else:
            bv[v]["a"] += 1
            bv[v]["ua"] += r["M"] or 0

    top = sorted(bv.items(), key=lambda x: -x[1]["t"])[:15]
    tbl = []
    for v, s in top:
        tc = s["c"] / s["t"] if s["t"] else 0
        tbl.append((v, s["t"], s["c"], s["a"], tc, s["uc"], s["ua"]))

    end = _rows(ws, dr, tbl, [None, INT, INT, INT, PCT, USD, USD])

    cats = Reference(ws, min_col=1, min_row=dr, max_row=end - 1)
    vc = Reference(ws, min_col=3, min_row=3, max_row=end - 1)
    va = Reference(ws, min_col=4, min_row=3, max_row=end - 1)

    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    bar.style = 10
    bar.title = "Tasa de Cierre por Vendedor"
    bar.y_axis.title = "Cotizaciones"
    bar.y_axis.numFmt = INT
    bar.width = CW
    bar.height = CH
    bar.add_data(vc, titles_from_data=True)
    bar.add_data(va, titles_from_data=True)
    bar.set_categories(cats)

    line = LineChart()
    vt = Reference(ws, min_col=5, min_row=3, max_row=end - 1)
    line.add_data(vt, titles_from_data=True)
    line.y_axis.title = "Tasa Cierre %"
    line.y_axis.numFmt = PCT

    _fix_combo(bar, line, rotate_x=-45)
    ws.add_chart(bar, f"A{end + 2}")


# ── Main ─────────────────────────────────────────────────────────────

def generate_dashboard(data_rows, serv_rows, fx, output_path):
    """Generate Dashboard_Insights.xlsx with 10 analysis sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    valid = [r for r in data_rows if (r["M"] or 0) > 0]

    _s1(wb, valid)
    _s2(wb, valid)
    _s3(wb, valid)
    _s4(wb, valid)
    _s5(wb, valid)
    _s6(wb, valid)
    _s7(wb, serv_rows, fx)
    _s8(wb, valid)
    _s9(wb, valid)
    _s10(wb, valid)

    wb.save(output_path)
    wb.close()
