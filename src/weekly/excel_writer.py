"""
Copia la plantilla Excel, escribe hojas DATA y DATA SERV, actualiza pivots.
"""

import shutil
from datetime import datetime

from openpyxl import load_workbook


def generate_weekly_excel(template, output, data_rows, serv_rows, serv_count, fx):
    """Copy template and populate DATA and DATA SERV sheets.

    Args:
        template: Path to the Excel template.
        output: Path for the output Excel file.
        data_rows: List of dicts for DATA sheet (combined SL + LLC).
        serv_rows: List of dicts for DATA SERV sheet (combined SL + LLC).
        serv_count: Number of service detail rows.
        fx: Exchange rates dict.

    Returns:
        The workbook object (still open for booking window writing).
    """
    shutil.copy2(template, output)
    wb = load_workbook(output)

    # ── Write DATA sheet ──────────────────────────────────────────────
    print(f"  Escribiendo hoja DATA ({len(data_rows)} filas)...")
    ws_data = wb["DATA"]

    # Clear existing data rows (keep header in row 1)
    for row in range(2, ws_data.max_row + 1):
        for col in range(1, 28):  # A-AA
            ws_data.cell(row, col).value = None

    col_map = {
        "A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7,
        "H": 8, "I": 9, "J": 10, "K": 11, "L": 12, "M": 13, "N": 14,
        "O": 15, "P": 16, "Q": 17, "R": 18, "S": 19, "T": 20, "U": 21,
        "V": 22, "W": 23, "X": 24, "Z": 26,
    }

    for i, row_data in enumerate(data_rows):
        excel_row = i + 2
        for col_letter, value in row_data.items():
            if col_letter in col_map and value is not None:
                ws_data.cell(excel_row, col_map[col_letter]).value = value

    # Format cells to match the template
    for row in range(2, len(data_rows) + 2):
        for col in [4, 5, 6, 22]:  # D, E, F, V (dates)
            cell = ws_data.cell(row, col)
            if cell.value and isinstance(cell.value, datetime):
                cell.number_format = "MM-DD-YY"
        for col in [10, 12, 13, 14, 15, 16]:  # J, L, M, N, O, P (numbers)
            cell = ws_data.cell(row, col)
            if cell.value is not None:
                cell.number_format = '0'
        cell = ws_data.cell(row, 17)  # Q (percentage)
        if cell.value is not None:
            cell.number_format = '0.00%'

    # ── Write DATA SERV sheet ─────────────────────────────────────────
    print(f"  Escribiendo hoja DATA SERV ({serv_count} filas)...")
    ws_serv = wb["DATA SERV"]

    # Clear existing data in columns A-V (preserve lookup tables Y-BO)
    for row in range(2, ws_serv.max_row + 1):
        for col in range(1, 23):  # A-V
            ws_serv.cell(row, col).value = None

    serv_col_map = {
        "B": 2, "C": 3, "E": 5, "F": 6, "G": 7,
        "H": 8, "I": 9, "K": 11, "L": 12, "O": 15,
    }

    for i, row_data in enumerate(serv_rows):
        excel_row = i + 2

        # Write static values
        for col_letter, value in row_data.items():
            if col_letter in serv_col_map and value is not None:
                ws_serv.cell(excel_row, serv_col_map[col_letter]).value = value

        # Determine company for entity-specific lookups
        company = row_data.get("B", "SL")
        r = excel_row

        # A: Oficina = XLOOKUP(vendedor -> pais de trabajo) — common table
        ws_serv.cell(r, 1).value = f'=_xlfn.XLOOKUP(D{r},BN:BN,BO:BO)'

        # D: vendedor = XLOOKUP(folio -> vendedor) — different per entity
        if company == "SL":
            ws_serv.cell(r, 4).value = f'=_xlfn.XLOOKUP(C{r},BC:BC,BD:BD)'
        else:  # LLC
            ws_serv.cell(r, 4).value = f'=_xlfn.XLOOKUP(C{r},BH:BH,BI:BI)'

        # J: tipo servicio = XLOOKUP(code -> significado) — common table
        ws_serv.cell(r, 10).value = f'=_xlfn.XLOOKUP(I{r},AS:AS,AU:AU)'

        # M: Subtotal_EUR = subtotal * Fx EUR
        ws_serv.cell(r, 13).value = f'=L{r}*_xlfn.XLOOKUP(K{r},AM:AM,AN:AN)'
        # N: Subtotal_USD = subtotal * Fx USD
        ws_serv.cell(r, 14).value = f'=L{r}*_xlfn.XLOOKUP(K{r},AM:AM,AO:AO)'
        # P: Monto Comision EUR
        ws_serv.cell(r, 16).value = f'=O{r}*_xlfn.XLOOKUP(K{r},AM:AM,AN:AN)'
        # Q: Monto Comision USD
        ws_serv.cell(r, 17).value = f'=O{r}*_xlfn.XLOOKUP(K{r},AM:AM,AO:AO)'

        # R: prov_nombre = XLOOKUP(proveedor -> nombre) — different per entity
        if company == "SL":
            ws_serv.cell(r, 18).value = f'=_xlfn.XLOOKUP(E{r},Y:Y,Z:Z)'
        else:  # LLC
            ws_serv.cell(r, 18).value = f'=_xlfn.XLOOKUP(E{r},AF:AF,AG:AG)'

        # S: Pais = XLOOKUP(proveedor -> pais) — different per entity
        if company == "SL":
            ws_serv.cell(r, 19).value = f'=_xlfn.XLOOKUP(E{r},Y:Y,AB:AB)'
        else:  # LLC
            ws_serv.cell(r, 19).value = f'=_xlfn.XLOOKUP(E{r},AF:AF,AI:AI)'

        # T: Semana = WEEKNUM(inicio_estancia)
        ws_serv.cell(r, 20).value = f'=WEEKNUM(G{r})'
        # U: Ano = YEAR(inicio_estancia)
        ws_serv.cell(r, 21).value = f'=YEAR(G{r})'
        # V: Rentabilidad = IFERROR(Q/N, "0")
        ws_serv.cell(r, 22).value = f'=IFERROR(Q{r}/N{r},"0")'

    # Format dates and numbers in DATA SERV
    for row in range(2, serv_count + 2):
        for col in [7, 8]:  # G, H (dates)
            cell = ws_serv.cell(row, col)
            if cell.value and isinstance(cell.value, datetime):
                cell.number_format = "MM-DD-YY"
        for col in [12, 13, 14, 15, 16, 17]:  # L-Q (numbers)
            cell = ws_serv.cell(row, col)
            if cell.value is not None and not isinstance(cell.value, str):
                cell.number_format = '#,##0.00'

    # ── Update FX rates in lookup table (AM-AO) ──────────────────────
    print(f"  Actualizando tipos de cambio en lookup table...")
    fx_rows = {
        4: "EUR", 5: "USD", 6: "CHF", 7: "GBP", 8: "GPB", 9: "JPY"
    }
    for row_num, currency in fx_rows.items():
        if currency in fx:
            ws_serv.cell(row_num, 40).value = fx[currency]["EUR"]  # AN
            ws_serv.cell(row_num, 41).value = fx[currency]["USD"]  # AO

    # Add MXN if not present
    if "MXN" in fx:
        for check_row in range(4, 15):
            if ws_serv.cell(check_row, 39).value == "MXN":
                ws_serv.cell(check_row, 40).value = fx["MXN"]["EUR"]
                ws_serv.cell(check_row, 41).value = fx["MXN"]["USD"]
                break
            if ws_serv.cell(check_row, 39).value is None:
                ws_serv.cell(check_row, 39).value = "MXN"
                ws_serv.cell(check_row, 40).value = fx["MXN"]["EUR"]
                ws_serv.cell(check_row, 41).value = fx["MXN"]["USD"]
                break

    # ── Update Pivot Table refresh settings ───────────────────────────
    print(f"  Configurando pivot tables para auto-refresh...")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, '_pivots') and ws._pivots:
            for pivot in ws._pivots:
                if pivot.cache and pivot.cache.cacheSource:
                    src = pivot.cache.cacheSource
                    if src.worksheetSource:
                        sheet_ref = src.worksheetSource.sheet
                        if sheet_ref == "DATA":
                            new_last_row = len(data_rows) + 1
                            src.worksheetSource.ref = f"A1:X{new_last_row}"
                        elif sheet_ref == "DATA SERV":
                            new_last_row = serv_count + 1
                            src.worksheetSource.ref = f"A1:V{new_last_row}"
                pivot.cache.refreshOnLoad = True

    return wb
