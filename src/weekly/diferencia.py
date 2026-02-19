"""
Escribe fórmulas de "Diferencia %" en la hoja "Weekly SL y LLC".

Fórmula por semana: =(B{row}-F{row})/F{row}
(Suma de Total Venta USD actual vs Venta año anterior USD)
Media al final: =AVERAGE(rango de diferencias)
"""

from openpyxl.styles import Font, PatternFill


def write_diferencia_pct(wb, data_rows):
    """
    Escribe fórmulas Excel de Diferencia % en la hoja "Weekly SL y LLC".

    Layout en columna I:
      - hdr_row:     "Diferencia %" (header)
      - hdr_row + 1: vacía
      - hdr_row + 2 en adelante: fórmula por semana
      - después de la última semana: AVERAGE de las diferencias

    Args:
        wb: Workbook abierto de openpyxl.
        data_rows: No se usa (mantenido por compatibilidad).

    Returns:
        Número de semanas procesadas, o 0 si la hoja no existe.
    """
    if "Weekly SL y LLC" not in wb.sheetnames:
        return 0

    ws = wb["Weekly SL y LLC"]

    # Encontrar fila de headers ("Semana" en col A)
    hdr_row = None
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() == "Semana":
            hdr_row = row
            break

    if not hdr_row:
        return 0

    # Header "Diferencia %"
    ws.cell(hdr_row, 9).value = "Diferencia %"
    ws.cell(hdr_row, 9).font = Font(bold=True)

    # Encontrar fila de Total general
    total_row = None
    for row in range(hdr_row + 1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip() == "Total general":
            total_row = row
            break

    # Limpiar hdr_row + 1 (puede tener fórmula del template)
    ws.cell(hdr_row + 1, 9).value = None

    # Escribir fórmulas desde hdr_row + 2
    end_row = total_row if total_row else ws.max_row + 1
    first_formula_row = None
    last_formula_row = None
    count = 0

    for row in range(hdr_row + 2, end_row):
        semana = ws.cell(row, 1).value
        if semana is None:
            continue

        cell = ws.cell(row, 9)
        cell.value = f"=IF(F{row}=0,0,(B{row}-F{row})/F{row})"
        cell.number_format = '0%'
        count += 1

        if first_formula_row is None:
            first_formula_row = row
        last_formula_row = row

    # Media de las diferencias en la fila siguiente a la última semana
    if first_formula_row and last_formula_row:
        avg_row = last_formula_row + 1
        cell_avg = ws.cell(avg_row, 9)
        cell_avg.value = f"=AVERAGE(I{first_formula_row}:I{last_formula_row})"
        cell_avg.number_format = '0.0%'
        cell_avg.font = Font(bold=True)
        cell_avg.fill = PatternFill(
            start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
        )

    return count
