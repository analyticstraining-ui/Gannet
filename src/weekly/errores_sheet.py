"""
Genera la hoja "ERRORES" con los posibles errores detectados en los datos.
"""

from openpyxl.styles import Font, PatternFill, Alignment


def write_errores_sheet(wb, errors):
    """
    Crea la hoja ERRORES con los errores detectados.

    Args:
        wb: Workbook abierto de openpyxl.
        errors: Lista de dicts con keys: Compañia, Folio, Error, Vendedor, Fecha.

    Returns:
        Número de errores escritos.
    """
    if not errors:
        return 0

    if "ERRORES" in wb.sheetnames:
        del wb["ERRORES"]
    ws = wb.create_sheet("ERRORES")

    hdr_font = Font(bold=True, color="FFFFFF")
    err_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    err_headers = ["Compañia", "Folio", "Error detectado", "Vendedor", "Fecha de registro"]
    err_widths = [12, 10, 50, 16, 18]
    for c, (h, w) in enumerate(zip(err_headers, err_widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = hdr_font
        cell.fill = err_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + c)].width = w

    for i, err in enumerate(errors, 2):
        ws.cell(i, 1, err.get("Compañia", ""))
        ws.cell(i, 2, err.get("Folio"))
        ws.cell(i, 3, err.get("Error", ""))
        ws.cell(i, 4, err.get("Vendedor", ""))
        fecha_val = err.get("Fecha")
        if fecha_val is not None:
            ws.cell(i, 5, fecha_val).number_format = "DD/MM/YY"
        else:
            ws.cell(i, 5, "")

    # Auto-filtro para toda la tabla
    last_row = len(errors) + 1
    ws.auto_filter.ref = f"A1:E{last_row}"

    return len(errors)
