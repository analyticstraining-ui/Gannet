"""
Genera la hoja "FX RATES" con tasas diarias del mes actual (datos BCE).
"""

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.fx_rates import get_current_month_daily_rates

# Monedas a mostrar (sin EUR=1.0, sin GPB=duplicado de GBP)
_CURRENCIES = ["USD", "GBP", "CHF", "JPY", "MXN"]


def write_fx_sheet(wb):
    """
    Crea la hoja FX RATES con tasas diarias del mes actual.

    Args:
        wb: Workbook abierto de openpyxl.

    Returns:
        (month_label, n_days): etiqueta del mes y cantidad de días escritos.
    """
    if "FX RATES" in wb.sheetnames:
        del wb["FX RATES"]
    ws = wb.create_sheet("FX RATES")

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    # Headers: Fecha | USD→EUR | USD→USD | GBP→EUR | ...
    headers = ["Fecha"]
    for cur in _CURRENCIES:
        headers.append(f"{cur}→EUR")
        headers.append(f"{cur}→USD")
    ncols = len(headers)

    # Título del mes
    month_label, daily = get_current_month_daily_rates()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, f"Tipos de cambio BCE — {month_label}").font = Font(bold=True, size=12)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 14
    for c in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Headers (row 3)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    # Daily rates
    for i, entry in enumerate(daily, 4):
        ws.cell(i, 1, entry["date"]).number_format = "DD/MM/YYYY"
        col = 2
        for cur in _CURRENCIES:
            rates = entry["rates"].get(cur, {})
            cell_eur = ws.cell(i, col, rates.get("EUR", 0))
            cell_usd = ws.cell(i, col + 1, rates.get("USD", 0))
            cell_eur.number_format = '0.000000'
            cell_usd.number_format = '0.000000'
            col += 2

    return month_label, len(daily)
