"""
Genera el Reporte AP & AR (Cuentas por Pagar / Cobrar).

AP = Accounts Payable  → cuánto se debe a proveedores
AR = Accounts Receivable → cuánto deben los clientes

Flujo (siguiendo instrucciones manuales):
  1. Copiar template
  2. Leer reserva → tabla lookup (folio, total_proveedor/total_cliente, moneda, fecha_inicio)
  3. Leer pago_proveedor → "ya pagado" (todos los aplicados) + "venta directa" (solo VD)
  4. Leer pago_cliente → "ya cobrado" (todos los aplicados) + "venta directa" (solo VD)
  5. Para LLC: tabla 4ZP/monedero de pago_cliente
  6. Escribir AP DATA y AR DATA con fórmulas XLOOKUP
  7. Crear Alertas Pago, Flags, FX Rates
  8. Force refresh de pivots
"""

import os
import smtplib
import shutil
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.properties import CalcProperties

from config import AP_AR_TEMPLATE_PATH, MONTH_NAMES_ES
from src.data_loader import load_reserva
from src.weekly.fx_sheet import write_fx_sheet

# Códigos de "venta directa" por entidad
VENTA_DIRECTA = {
    "SL": {"00B"},
    "LLC": {"4E2", "E04"},
}

# Hojas a conservar en el output (nombres del template LLC)
_KEEP_SHEETS = {"AP DATA", "AR DATA", "AP Pivot", "AR Pivot"}

# Monedas para la tabla FX
_FX_CURRENCIES = ["EUR", "USD", "CHF", "GBP", "GPB", "JPY", "MXN"]


# ── Carga de CSVs ────────────────────────────────────────────────────────

def _find_csv(data_dir, base_name):
    """Busca un CSV por nombre, eligiendo el más reciente si hay variantes."""
    import glob
    exact = os.path.join(data_dir, f"{base_name}.csv")
    if os.path.exists(exact):
        candidates = [exact]
    else:
        candidates = []

    pattern = os.path.join(data_dir, f"{base_name}*.csv")
    candidates.extend(glob.glob(pattern))

    if not candidates:
        return None

    candidates = list(set(candidates))
    candidates.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return candidates[0]


def _load_payment_csv(data_dir, filename):
    """Lee un CSV de pagos (pago_proveedor o pago_cliente)."""
    path = _find_csv(data_dir, filename)
    if not path:
        raise FileNotFoundError(f"No se encuentra {filename}.csv en {data_dir}")
    print(f"    Leyendo {os.path.basename(path)}...")
    df = pd.read_csv(path, encoding="latin-1")
    print(f"    {len(df)} registros leídos.")
    return df


# ── Lookup de reserva ────────────────────────────────────────────────────

def _build_reserva_lookup(reserva_df):
    """Crea {folio: {moneda, fecha_inicio, total_proveedor, total_cliente}}."""
    lookup = {}
    for _, r in reserva_df.iterrows():
        folio = r["folio"]
        moneda = str(r.get("moneda", "EUR")).strip()
        fecha_inicio = r.get("fecha_inicio")
        if pd.notna(fecha_inicio):
            if isinstance(fecha_inicio, str):
                try:
                    fecha_inicio = pd.to_datetime(fecha_inicio)
                except Exception:
                    fecha_inicio = None
        else:
            fecha_inicio = None

        total_prov = r.get("total_proveedor", 0)
        if pd.isna(total_prov):
            total_prov = 0
        total_cli = r.get("total_cliente", 0)
        if pd.isna(total_cli):
            total_cli = 0

        lookup[folio] = {
            "moneda": moneda,
            "fecha_inicio": fecha_inicio,
            "total_proveedor": float(total_prov),
            "total_cliente": float(total_cli),
        }
    return lookup


# ── Construcción de datos ────────────────────────────────────────────────

def _build_ya_pagado(payment_df, date_col):
    """Pivot de TODOS los pagos aplicados (sin exclusión de venta directa).

    Filtros: cancelado=0, date_col != '0000-00-00'
    Returns {reserva: sum_monto}
    """
    applied = payment_df[
        (payment_df["cancelado"] == 0) &
        (payment_df[date_col].astype(str).str.strip() != "0000-00-00")
    ]
    return applied.groupby("reserva")["monto"].sum().to_dict()


def _build_venta_directa(payment_df, vd_codes, date_col, monto_col):
    """Pivot de SOLO pagos de venta directa aplicados.

    Filtros: cancelado=0, date_col != '0000-00-00', forma_pago IN vd_codes
    Returns {reserva: sum_monto}
    """
    filtered = payment_df[
        (payment_df["cancelado"] == 0) &
        (payment_df[date_col].astype(str).str.strip() != "0000-00-00") &
        (payment_df["forma_pago"].isin(vd_codes))
    ]
    return filtered.groupby("reserva")[monto_col].sum().to_dict()


def _build_moneda_raw(payment_df, date_col):
    """Lista de (reserva, moneda) de pagos aplicados, un row por reserva.

    Para tabla W-X (currency lookup).
    """
    applied = payment_df[
        (payment_df["cancelado"] == 0) &
        (payment_df[date_col].astype(str).str.strip() != "0000-00-00")
    ]
    # Un row por reserva (primera moneda encontrada)
    deduped = applied.drop_duplicates(subset="reserva")[["reserva", "moneda"]]
    return list(deduped.itertuples(index=False, name=None))


def _build_4zp_records(pago_cli_df):
    """Registros 4ZP/monedero de pago_cliente (solo LLC).

    Filtro: cancelado=0, forma_pago='4ZP'
    Returns list of (reserva, monto, moneda)
    """
    filtered = pago_cli_df[
        (pago_cli_df["cancelado"] == 0) &
        (pago_cli_df["forma_pago"] == "4ZP")
    ]
    return [
        (row["reserva"], row["monto"], str(row.get("moneda", "MXN")).strip())
        for _, row in filtered.iterrows()
    ]


# ── Escritura de hojas de datos ──────────────────────────────────────────
#
# AP DATA layout (siguiendo template LLC):
#   A:  reserva (literal)      F: ya pagado (literal)
#   B-E, G-N: fórmulas XLOOKUP
#   Q-R: venta directa pivot   S: moneda (XLOOKUP)  T: USD  U: forma_pago
#   W-X: reserva+moneda raw
#   Z-AD: reserva lookup (folio, total_prov, moneda, USD, fecha_inicio)
#   AG-AI: FX table
#
# AR DATA layout (siguiendo template LLC):
#   A-N: similar a AP
#   Q-U: venta directa pivot
#   W-X: reserva+moneda raw
#   Z-AD: tabla 4ZP (solo LLC)
#   AF-AI: reserva lookup (folio, moneda, total_cli, fecha_inicio)
#   AM-AO: FX table

def _clear_sheet(ws, max_col):
    """Limpia datos existentes preservando fila 1 (headers)."""
    for row in range(2, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def _write_fx_table(ws, fx, col_divisa, col_eur, col_usd, start_row=1):
    """Escribe la mini-tabla FX (DIVISA, Fx EUR, Fx USD)."""
    ws.cell(start_row, col_divisa).value = "DIVISA"
    ws.cell(start_row, col_eur).value = "Fx EUR"
    ws.cell(start_row, col_usd).value = "Fx USD"
    for i, curr in enumerate(_FX_CURRENCIES):
        r = start_row + 1 + i
        ws.cell(r, col_divisa).value = curr
        if curr in fx:
            ws.cell(r, col_eur).value = fx[curr]["EUR"]
            ws.cell(r, col_usd).value = fx[curr]["USD"]


def _write_ap_data(ws, ya_pagado, venta_directa, moneda_raw,
                   reserva_lookup, fx, vd_label):
    """Escribe AP DATA siguiendo el layout del template LLC."""
    _clear_sheet(ws, 35)  # A-AI

    folios = sorted(ya_pagado.keys())

    # ── A + F: reserva y ya pagado (literales) ──
    for i, folio in enumerate(folios):
        r = i + 2
        ws.cell(r, 1).value = folio                                    # A
        ws.cell(r, 6).value = round(ya_pagado[folio], 2)               # F

    # ── B-E, G-N: fórmulas XLOOKUP ──
    for i, folio in enumerate(folios):
        r = i + 2
        # B: moneda desde reserva lookup (Z:AB)
        ws.cell(r, 2).value = f'=_xlfn.XLOOKUP(A{r},Z:Z,AB:AB,"not found")'
        # C: total_proveedor - venta_directa
        ws.cell(r, 3).value = f'=_xlfn.XLOOKUP(A{r},Z:Z,AA:AA)-_xlfn.XLOOKUP(A{r},Q:Q,R:R,"0")'
        # D: C * FX EUR
        ws.cell(r, 4).value = f'=C{r}*_xlfn.XLOOKUP(B{r},AG:AG,AH:AH)'
        # E: C * FX USD
        ws.cell(r, 5).value = f'=C{r}*_xlfn.XLOOKUP(B{r},AG:AG,AI:AI)'
        # G: F * FX EUR
        ws.cell(r, 7).value = f'=F{r}*_xlfn.XLOOKUP(B{r},AG:AG,AH:AH)'
        # H: F * FX USD
        ws.cell(r, 8).value = f'=F{r}*_xlfn.XLOOKUP(B{r},AG:AG,AI:AI)'
        # I: restante = C - F + venta_directa
        ws.cell(r, 9).value = f'=C{r}-F{r}+_xlfn.XLOOKUP(A{r},Q:Q,R:R,"0")'
        # J: restante USD
        ws.cell(r, 10).value = f'=I{r}*_xlfn.XLOOKUP(B{r},AG:AG,AI:AI)'
        # K: fecha_inicio
        ws.cell(r, 11).value = f'=_xlfn.XLOOKUP(A{r},Z:Z,AD:AD)'
        # L: fecha - 30
        ws.cell(r, 12).value = f'=K{r}-30'
        # M: mes
        ws.cell(r, 13).value = f'=MONTH(L{r})'
        # N: año
        ws.cell(r, 14).value = f'=YEAR(L{r})'

    # ── Q-U: venta directa pivot ──
    vd_folios = sorted(venta_directa.keys())
    for i, folio in enumerate(vd_folios):
        r = i + 2
        ws.cell(r, 17).value = folio                                    # Q
        ws.cell(r, 18).value = round(venta_directa[folio], 2)           # R
        # S: moneda via XLOOKUP desde W-X
        ws.cell(r, 19).value = f'=_xlfn.XLOOKUP(Q{r},W:W,X:X)'
        # T: monto USD
        ws.cell(r, 20).value = f'=R{r}*_xlfn.XLOOKUP(S{r},AG:AG,AI:AI)'
        ws.cell(r, 21).value = vd_label                                 # U

    # ── W-X: reserva + moneda raw ──
    for i, (reserva, moneda) in enumerate(moneda_raw):
        r = i + 2
        ws.cell(r, 23).value = reserva                                  # W
        ws.cell(r, 24).value = str(moneda).strip() if pd.notna(moneda) else "EUR"  # X

    # ── Z-AD: reserva lookup (folio, total_proveedor, moneda, USD, fecha_inicio) ──
    all_folios = sorted(reserva_lookup.keys())
    for i, folio in enumerate(all_folios):
        r = i + 2
        info = reserva_lookup[folio]
        ws.cell(r, 26).value = folio                                     # Z
        ws.cell(r, 27).value = round(info["total_proveedor"], 2)         # AA
        ws.cell(r, 28).value = info["moneda"]                            # AB
        # AC: total_proveedor USD
        ws.cell(r, 29).value = f'=AA{r}*_xlfn.XLOOKUP(AB{r},AG:AG,AI:AI)'
        fi = info["fecha_inicio"]
        if fi is not None:
            ws.cell(r, 30).value = fi                                    # AD
            ws.cell(r, 30).number_format = 'DD/MM/YYYY'

    # ── AG-AI: FX table ──
    _write_fx_table(ws, fx, col_divisa=33, col_eur=34, col_usd=35)

    return len(folios)


def _write_ar_data(ws, ya_pagado, venta_directa, moneda_raw,
                   reserva_lookup, zp_records, fx, company, vd_label):
    """Escribe AR DATA siguiendo el layout del template LLC."""
    _clear_sheet(ws, 41)  # A-AO

    folios = sorted(ya_pagado.keys())

    # ── A + F: reserva y ya cobrado (literales) ──
    for i, folio in enumerate(folios):
        r = i + 2
        ws.cell(r, 1).value = folio                                    # A
        ws.cell(r, 6).value = round(ya_pagado[folio], 2)               # F

    # ── B-E, G-N: fórmulas XLOOKUP ──
    for i, folio in enumerate(folios):
        r = i + 2
        # B: moneda desde reserva lookup (AF:AG)
        ws.cell(r, 2).value = f'=_xlfn.XLOOKUP(A{r},AF:AF,AG:AG,"not found")'
        # C: total_cliente - venta_directa
        ws.cell(r, 3).value = f'=_xlfn.XLOOKUP(A{r},AF:AF,AH:AH)-_xlfn.XLOOKUP(A{r},Q:Q,R:R,"0")'
        # D: C * FX EUR
        ws.cell(r, 4).value = f'=C{r}*_xlfn.XLOOKUP(B{r},AM:AM,AN:AN)'
        # E: C * FX USD
        ws.cell(r, 5).value = f'=C{r}*_xlfn.XLOOKUP(B{r},AM:AM,AO:AO)'
        # G: F * FX EUR
        ws.cell(r, 7).value = f'=F{r}*_xlfn.XLOOKUP(B{r},AM:AM,AN:AN)'
        # H: F * FX USD
        ws.cell(r, 8).value = f'=F{r}*_xlfn.XLOOKUP(B{r},AM:AM,AO:AO)'
        # I: restante = C - F + venta_directa [- 4ZP para LLC]
        if company == "LLC":
            ws.cell(r, 9).value = (
                f'=C{r}-F{r}+_xlfn.XLOOKUP(A{r},Q:Q,R:R,"0")'
                f'-_xlfn.XLOOKUP(A{r},Z:Z,AA:AA,"0")'
            )
        else:
            ws.cell(r, 9).value = f'=C{r}-F{r}+_xlfn.XLOOKUP(A{r},Q:Q,R:R,"0")'
        # J: restante USD
        ws.cell(r, 10).value = f'=I{r}*_xlfn.XLOOKUP(B{r},AM:AM,AO:AO)'
        # K: fecha_inicio
        ws.cell(r, 11).value = f'=_xlfn.XLOOKUP(A{r},AF:AF,AI:AI)'
        # L: fecha - 30
        ws.cell(r, 12).value = f'=K{r}-30'
        # M: mes
        ws.cell(r, 13).value = f'=MONTH(L{r})'
        # N: año
        ws.cell(r, 14).value = f'=YEAR(L{r})'

    # ── Q-U: venta directa pivot ──
    vd_folios = sorted(venta_directa.keys())
    for i, folio in enumerate(vd_folios):
        r = i + 2
        ws.cell(r, 17).value = folio                                    # Q
        ws.cell(r, 18).value = round(venta_directa[folio], 2)           # R
        # S: moneda via XLOOKUP desde W-X
        ws.cell(r, 19).value = f'=_xlfn.XLOOKUP(Q{r},W:W,X:X)'
        # T: monto USD
        ws.cell(r, 20).value = f'=R{r}*_xlfn.XLOOKUP(S{r},AM:AM,AO:AO)'
        ws.cell(r, 21).value = vd_label                                 # U

    # ── W-X: reserva + moneda raw ──
    for i, (reserva, moneda) in enumerate(moneda_raw):
        r = i + 2
        ws.cell(r, 23).value = reserva                                  # W
        ws.cell(r, 24).value = str(moneda).strip() if pd.notna(moneda) else "EUR"  # X

    # ── Z-AD: tabla 4ZP/monedero (solo LLC) ──
    if company == "LLC" and zp_records:
        # Headers
        ws.cell(1, 26).value = "reserva"
        ws.cell(1, 27).value = "monto"
        ws.cell(1, 28).value = "moneda"
        ws.cell(1, 29).value = "monto USD"
        ws.cell(1, 30).value = "forma_pago"
        for i, (reserva, monto, moneda) in enumerate(zp_records):
            r = i + 2
            ws.cell(r, 26).value = reserva                              # Z
            ws.cell(r, 27).value = round(monto, 2)                      # AA
            ws.cell(r, 28).value = moneda                                # AB
            ws.cell(r, 29).value = f'=AA{r}*_xlfn.XLOOKUP(AB{r},AM:AM,AO:AO)'  # AC
            ws.cell(r, 30).value = "4ZP"                                 # AD

    # ── AF-AI: reserva lookup (folio, moneda, total_cliente, fecha_inicio) ──
    all_folios = sorted(reserva_lookup.keys())
    for i, folio in enumerate(all_folios):
        r = i + 2
        info = reserva_lookup[folio]
        ws.cell(r, 32).value = folio                                     # AF
        ws.cell(r, 33).value = info["moneda"]                            # AG
        ws.cell(r, 34).value = round(info["total_cliente"], 2)           # AH
        fi = info["fecha_inicio"]
        if fi is not None:
            ws.cell(r, 35).value = fi                                    # AI
            ws.cell(r, 35).number_format = 'DD/MM/YYYY'

    # ── AM-AO: FX table ──
    _write_fx_table(ws, fx, col_divisa=39, col_eur=40, col_usd=41)

    return len(folios)


# ── Alertas de pago ─────────────────────────────────────────────────────

def _load_proveedor_lookup(data_dir):
    """Carga proveedor.csv y crea {clave: {nombre, email, ciudad}}."""
    path = _find_csv(data_dir, "proveedor")
    if not path:
        return {}
    df = pd.read_csv(path, encoding="latin-1")
    lookup = {}
    for _, r in df.iterrows():
        clave = r.get("clave")
        if pd.isna(clave):
            continue
        lookup[int(clave)] = {
            "nombre": str(r.get("nombre", "")).strip() if pd.notna(r.get("nombre")) else "",
            "email": str(r.get("correo_e_contacto", "")).strip() if pd.notna(r.get("correo_e_contacto")) else "",
            "ciudad": str(r.get("ciudad", "")).strip() if pd.notna(r.get("ciudad")) else "",
        }
    return lookup


def _build_alertas_data(pago_prov_df, prov_lookup):
    """Identifica pagos a proveedores sin aplicar, próximos a vencer o ya excedidos."""
    today = datetime.now().date()

    unpaid = pago_prov_df[
        (pago_prov_df["fecha_aplicacion"].astype(str).str.strip() == "0000-00-00") &
        (pago_prov_df["monto_monedero"] == 0)
    ].copy()

    if unpaid.empty:
        return [], []

    unpaid["fecha_limite_dt"] = pd.to_datetime(unpaid["fecha_limite"], errors="coerce")

    proximos = []
    excedidos = []

    for _, row in unpaid.iterrows():
        fl = row["fecha_limite_dt"]
        if pd.isna(fl):
            continue
        fl_date = fl.date()
        days_remaining = (fl_date - today).days

        if days_remaining > 7:
            continue

        prov_code = row.get("proveedor", "")
        prov_info = prov_lookup.get(int(prov_code), {}) if pd.notna(prov_code) else {}

        record = {
            "reserva": row["reserva"],
            "vendedor": str(row.get("vendedor", "")),
            "fecha": str(row.get("fecha", "")),
            "proveedor_code": prov_code,
            "proveedor_nombre": prov_info.get("nombre", ""),
            "proveedor_email": prov_info.get("email", ""),
            "proveedor_ciudad": prov_info.get("ciudad", ""),
            "fecha_limite": fl_date,
            "monto": row.get("monto", 0),
        }

        if days_remaining < 0:
            excedidos.append(record)
        else:
            proximos.append(record)

    proximos.sort(key=lambda x: x["fecha_limite"])
    excedidos.sort(key=lambda x: x["fecha_limite"])

    return proximos, excedidos


def _write_alertas_sheet(wb, proximos, excedidos):
    """Crea la hoja Alertas Pago con dos tablas: próximos a vencer y ya excedidos."""
    ws = wb.create_sheet("Alertas Pago")

    title_font_white = Font(bold=True, size=13, color="FFFFFF")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    orange_fill = PatternFill("solid", fgColor="FF8C00")
    red_fill = PatternFill("solid", fgColor="CC0000")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    cols = ["Reserva", "Vendedor", "Fecha", "Proveedor", "Nombre Proveedor",
            "Email Contacto", "Ciudad", "Fecha Límite", "Monto"]
    keys = ["reserva", "vendedor", "fecha", "proveedor_code", "proveedor_nombre",
            "proveedor_email", "proveedor_ciudad", "fecha_limite", "monto"]

    def write_table(start_row, title, fill, records):
        ws.cell(start_row, 1).value = title
        ws.cell(start_row, 1).font = title_font_white
        for c in range(1, len(cols) + 1):
            ws.cell(start_row, c).fill = fill
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=len(cols))

        hr = start_row + 1
        for c, name in enumerate(cols, 1):
            cell = ws.cell(hr, c)
            cell.value = name
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        if not records:
            ws.cell(hr + 1, 1).value = "Sin registros"
            return hr + 2

        for i, rec in enumerate(records):
            r = hr + 1 + i
            for c, key in enumerate(keys, 1):
                cell = ws.cell(r, c)
                cell.value = rec.get(key, "")
                cell.border = thin_border
                if key == "fecha_limite":
                    cell.number_format = 'DD/MM/YYYY'
                elif key == "monto":
                    cell.number_format = '#,##0.00'

        return hr + 1 + len(records)

    end_row = write_table(
        1, f"PROXIMOS A VENCER (7 dias o menos) - {len(proximos)} registros",
        orange_fill, proximos,
    )

    write_table(
        end_row + 2, f"YA EXCEDIDOS - {len(excedidos)} registros",
        red_fill, excedidos,
    )

    widths = {"A": 12, "B": 15, "C": 14, "D": 12, "E": 28,
              "F": 30, "G": 16, "H": 16, "I": 14}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    return len(proximos), len(excedidos)


# ── Email de alertas (solo Madrid / SL) ─────────────────────────────────

_ALERTA_EMAIL_TO = "analytics.training@gannetworld.com"


def _send_alertas_email(proximos, entity_label):
    """Envía un email con la tabla de pagos próximos a vencer (solo Madrid)."""
    smtp_host = os.environ.get("SMTP_HOST", "")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")

    if not all([smtp_host, smtp_user, smtp_pass]):
        print("  Email: variables SMTP no configuradas (SMTP_HOST, SMTP_USER, SMTP_PASS)")
        print("  Saltando envío de email.")
        return False

    today_str = datetime.now().strftime("%d/%m/%Y")

    rows_html = ""
    for rec in proximos:
        fl = rec["fecha_limite"]
        fl_str = fl.strftime("%d/%m/%Y") if hasattr(fl, "strftime") else str(fl)
        rows_html += f"""<tr>
            <td>{rec['reserva']}</td>
            <td>{rec['vendedor']}</td>
            <td>{rec['fecha']}</td>
            <td>{rec['proveedor_code']}</td>
            <td>{rec['proveedor_nombre']}</td>
            <td>{rec['proveedor_email']}</td>
            <td>{rec['proveedor_ciudad']}</td>
            <td>{fl_str}</td>
            <td style="text-align:right">{rec['monto']:,.2f}</td>
        </tr>"""

    html = f"""<html><body>
    <h2>Alerta Pagos Proveedores - {entity_label}</h2>
    <p>Fecha de ejecucion: {today_str}</p>
    <p>Hay <b>{len(proximos)}</b> pagos proximos a vencer (7 dias o menos):</p>
    <table border="1" cellpadding="6" cellspacing="0"
           style="border-collapse:collapse; font-family:Arial; font-size:12px;">
      <tr style="background:#4472C4; color:white;">
        <th>Reserva</th><th>Vendedor</th><th>Fecha</th>
        <th>Proveedor</th><th>Nombre</th><th>Email</th><th>Ciudad</th>
        <th>Fecha Limite</th><th>Monto</th>
      </tr>
      {rows_html}
    </table>
    <br><p style="color:gray; font-size:11px;">Generado automaticamente por Gannet Reports.</p>
    </body></html>"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"Alerta Pagos Proveedores {entity_label} - {today_str}"
    msg["From"] = smtp_user
    msg["To"] = _ALERTA_EMAIL_TO
    msg.attach(MIMEText(html, "html"))

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, [_ALERTA_EMAIL_TO], msg.as_string())
        print(f"  Email enviado a {_ALERTA_EMAIL_TO}")
        return True
    except Exception as e:
        print(f"  Email falló: {e}")
        return False


# ── Flags (anomalías) ────────────────────────────────────────────────────

def _get_fx_usd(moneda, fx):
    """Obtiene tasa moneda→USD desde fx dict."""
    rate = fx.get(moneda, {}).get("USD")
    if rate:
        return rate
    from config import FALLBACK_FX
    return FALLBACK_FX.get(moneda, {}).get("USD", 1)


def _build_flags_data(ap_ya_pagado, ar_ya_pagado, ap_venta_directa, ar_venta_directa,
                      reserva_lookup, fx, zp_records=None):
    """Detecta anomalías usando la misma lógica que las fórmulas del template.

    AP restante = total_proveedor - ya_pagado → si < 0, pagado de más
    AR restante = total_cliente - ya_cobrado → si < 0, cobrado de más
    """
    # Build 4ZP lookup
    zp_by_folio = {}
    if zp_records:
        for reserva, monto, moneda in zp_records:
            zp_by_folio[reserva] = zp_by_folio.get(reserva, 0) + monto

    ap_flags = []
    for folio in ap_ya_pagado:
        info = reserva_lookup.get(folio, {})
        total = info.get("total_proveedor", 0)
        pagado = ap_ya_pagado[folio]
        restante = total - pagado
        moneda = info.get("moneda", "EUR")
        fx_usd = _get_fx_usd(moneda, fx)
        restante_usd = round(restante * fx_usd, 2)
        if restante_usd < 0:
            ap_flags.append({
                "folio": folio,
                "moneda": moneda,
                "total": round(total, 2),
                "pagado": round(pagado, 2),
                "restante": round(restante, 2),
                "restante_usd": restante_usd,
            })

    ar_flags = []
    for folio in ar_ya_pagado:
        info = reserva_lookup.get(folio, {})
        total = info.get("total_cliente", 0)
        cobrado = ar_ya_pagado[folio]
        zp_monto = zp_by_folio.get(folio, 0)
        restante = total - cobrado - zp_monto
        moneda = info.get("moneda", "EUR")
        fx_usd = _get_fx_usd(moneda, fx)
        restante_usd = round(restante * fx_usd, 2)
        if restante_usd < 0:
            ar_flags.append({
                "folio": folio,
                "moneda": moneda,
                "total": round(total, 2),
                "pagado": round(cobrado, 2),
                "restante": round(restante, 2),
                "restante_usd": restante_usd,
            })

    ap_flags.sort(key=lambda x: x["restante_usd"])
    ar_flags.sort(key=lambda x: x["restante_usd"])
    return ap_flags, ar_flags


def _write_flags_sheet(wb, ap_flags, ar_flags):
    """Crea la hoja Flags con dos tablas."""
    ws = wb.create_sheet("Flags")

    title_font = Font(bold=True, size=13, color="FFFFFF")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    red_fill = PatternFill("solid", fgColor="CC0000")
    orange_fill = PatternFill("solid", fgColor="FF8C00")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ap_cols = ["Folio", "Moneda", "Total Comprometido", "Ya Pagado",
               "Restante", "Restante USD"]
    ar_cols = ["Folio", "Moneda", "Total Cliente", "Ya Cobrado",
               "Restante", "Restante USD"]
    keys = ["folio", "moneda", "total", "pagado", "restante", "restante_usd"]

    def write_table(start_row, title, fill, cols, records):
        ws.cell(start_row, 1).value = title
        ws.cell(start_row, 1).font = title_font
        for c in range(1, len(cols) + 1):
            ws.cell(start_row, c).fill = fill
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=len(cols))

        hr = start_row + 1
        for c, name in enumerate(cols, 1):
            cell = ws.cell(hr, c)
            cell.value = name
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        if not records:
            ws.cell(hr + 1, 1).value = "Sin registros"
            return hr + 2

        for i, rec in enumerate(records):
            r = hr + 1 + i
            for c, key in enumerate(keys, 1):
                cell = ws.cell(r, c)
                cell.value = rec.get(key, "")
                cell.border = thin_border
                if key in ("total", "pagado", "restante", "restante_usd"):
                    cell.number_format = "#,##0.00"

        return hr + 1 + len(records)

    end_row = write_table(
        1, f"PROVEEDORES PAGADOS DE MAS - {len(ap_flags)} registros",
        red_fill, ap_cols, ap_flags,
    )

    write_table(
        end_row + 2, f"CLIENTES PAGADO DE MENOS - {len(ar_flags)} registros",
        orange_fill, ar_cols, ar_flags,
    )

    widths = {"A": 12, "B": 10, "C": 20, "D": 16, "E": 14, "F": 16}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    return len(ap_flags), len(ar_flags)


# ── Limpieza y pivot refresh ─────────────────────────────────────────────

def _cleanup_sheets(wb):
    """Elimina hojas innecesarias del template."""
    to_delete = [sn for sn in wb.sheetnames if sn not in _KEEP_SHEETS]
    for sn in to_delete:
        del wb[sn]


def _force_pivot_refresh(wb):
    """Fuerza el recálculo de pivots y fórmulas al abrir."""
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True

    seen = set()
    for ws in wb.worksheets:
        for pt in getattr(ws, '_pivots', []):
            cache = pt.cache
            cid = id(cache)
            if cid in seen:
                continue
            seen.add(cid)
            cache.refreshOnLoad = True


# ── Función principal ────────────────────────────────────────────────────

def generate_ap_ar_report(data_dir, output_dir, fx, company, entity_label,
                          year=None, month=None):
    """Genera el Reporte AP & AR para una entidad."""
    today = datetime.now()
    year = year or today.year
    month = month or today.month
    month_name = MONTH_NAMES_ES.get(month, str(month)).capitalize()

    # Verificar template
    if not os.path.exists(AP_AR_TEMPLATE_PATH):
        print(f"  ERROR: No se encuentra el template: {AP_AR_TEMPLATE_PATH}")
        return None

    vd_codes = VENTA_DIRECTA.get(company, set())
    vd_label = next(iter(vd_codes)) if len(vd_codes) == 1 else ",".join(sorted(vd_codes))
    print(f"  Entidad: {entity_label} ({company})")
    print(f"  Códigos venta directa: {vd_codes}")

    # 1. Cargar reserva (filtrada, no canceladas)
    reserva_df = load_reserva(data_dir)
    reserva_lookup = _build_reserva_lookup(reserva_df)
    print(f"  Reservas activas: {len(reserva_lookup)}")

    # 2. Cargar pagos y proveedores
    pago_prov_df = _load_payment_csv(data_dir, "pago_proveedor")
    pago_cli_df = _load_payment_csv(data_dir, "pago_cliente")
    prov_lookup = _load_proveedor_lookup(data_dir)
    print(f"  Proveedores cargados: {len(prov_lookup)}")

    # 3. AP: ya pagado (todos los aplicados, sin excluir VD)
    ap_ya_pagado = _build_ya_pagado(pago_prov_df, "fecha_aplicacion")
    print(f"  AP ya pagado: {len(ap_ya_pagado)} reservas")

    # 4. AP: venta directa (solo VD, monto_pagado)
    ap_venta_directa = _build_venta_directa(
        pago_prov_df, vd_codes, "fecha_aplicacion", "monto_pagado"
    )
    print(f"  AP venta directa: {len(ap_venta_directa)} reservas")

    # 5. AP: moneda raw
    ap_moneda_raw = _build_moneda_raw(pago_prov_df, "fecha_aplicacion")

    # 6. AR: ya cobrado (todos los aplicados, sin excluir VD)
    ar_ya_pagado = _build_ya_pagado(pago_cli_df, "fecha_proceso")
    print(f"  AR ya cobrado: {len(ar_ya_pagado)} reservas")

    # 7. AR: venta directa (solo VD, monto)
    ar_venta_directa = _build_venta_directa(
        pago_cli_df, vd_codes, "fecha_proceso", "monto"
    )
    print(f"  AR venta directa: {len(ar_venta_directa)} reservas")

    # 8. AR: moneda raw
    ar_moneda_raw = _build_moneda_raw(pago_cli_df, "fecha_proceso")

    # 9. AR: 4ZP/monedero (solo LLC)
    zp_records = []
    if company == "LLC":
        zp_records = _build_4zp_records(pago_cli_df)
        print(f"  AR 4ZP/monedero: {len(zp_records)} registros")

    # 10. Alertas de pago
    proximos, excedidos = _build_alertas_data(pago_prov_df, prov_lookup)
    print(f"  Alertas: {len(proximos)} próximos a vencer, {len(excedidos)} ya excedidos")

    # 11. Copiar template y abrir
    filename = f"Report_AP_&_AR_{entity_label}_{month_name}_{year}.xlsx"
    filepath = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)
    shutil.copy2(AP_AR_TEMPLATE_PATH, filepath)

    wb = load_workbook(filepath)

    # 12. Limpiar hojas innecesarias
    _cleanup_sheets(wb)
    print(f"  Hojas: {wb.sheetnames}")

    # 13. Escribir AP DATA
    ws_ap = wb["AP DATA"]
    n_ap = _write_ap_data(ws_ap, ap_ya_pagado, ap_venta_directa,
                          ap_moneda_raw, reserva_lookup, fx, vd_label)
    print(f"  AP DATA: {n_ap} filas escritas")

    # 14. Escribir AR DATA
    ws_ar = wb["AR DATA"]
    n_ar = _write_ar_data(ws_ar, ar_ya_pagado, ar_venta_directa,
                          ar_moneda_raw, reserva_lookup, zp_records,
                          fx, company, vd_label)
    print(f"  AR DATA: {n_ar} filas escritas")

    # 15. Alertas Pago
    n_prox, n_exc = _write_alertas_sheet(wb, proximos, excedidos)
    print(f"  Alertas Pago: {n_prox} próximos, {n_exc} excedidos")

    # 16. Flags (anomalías)
    ap_flags, ar_flags = _build_flags_data(
        ap_ya_pagado, ar_ya_pagado,
        ap_venta_directa, ar_venta_directa,
        reserva_lookup, fx, zp_records,
    )
    n_ap_flags, n_ar_flags = _write_flags_sheet(wb, ap_flags, ar_flags)
    print(f"  Flags: {n_ap_flags} proveedores pagados de más, {n_ar_flags} clientes pagados de menos")

    # 17. FX Rates (tasas diarias del mes)
    month_label, n_days = write_fx_sheet(wb)
    print(f"  FX Rates: {n_days} días de {month_label}")

    # 18. Force refresh (pivots + fórmulas)
    _force_pivot_refresh(wb)
    print(f"  Force refresh configurado")

    # 19. Guardar
    wb.save(filepath)
    wb.close()

    # 20. Email de alertas (solo Madrid/SL, solo si hay próximos)
    if company == "SL" and proximos:
        _send_alertas_email(proximos, entity_label)

    return filepath
