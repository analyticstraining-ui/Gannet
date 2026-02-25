"""
Genera el Reporte de Comisiones Pendientes Proveedores.

Muestra comisiones pendientes de pago a proveedores, agrupadas por entidad
(SL/Madrid y LLC/Mexico). Incluye:
  - Data SL / Data LLC: datos filtrados de dreserva con formulas XLOOKUP
  - Com Pend SL / Com Pend LLC: pivot de comision EUR por anio/mes (45d post checkout)
  - Desglose: tabla combinada con info decodificada de proveedores
  - FX Rates: tasas diarias del mes

Filtros aplicados a dreserva.csv:
  comision_pendiente = 1
  monto_comision != 0
  fecha_pago = '0000-00-00'
  servicio_cancelado = 0
"""

import os
import glob
import shutil
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

from config import COM_PEND_PROV_TEMPLATE_PATH, FALLBACK_FX, MONTH_NAMES_ES
from src.weekly.fx_sheet import write_fx_sheet

# Hojas a conservar del template
_KEEP_SHEETS = {"Data LLC", "Com Pend LLC", "Data SL", "Com Pend SL"}

# Monedas para la tabla FX (incluye GPB = typo en datos fuente)
_FX_CURRENCIES = ["EUR", "USD", "CHF", "GBP", "GPB", "JPY", "MXN"]

_ENTITY_LABELS = {"SL": "Madrid", "LLC": "Mexico"}


# ── Carga de CSVs ────────────────────────────────────────────────────────

def _find_csv(data_dir, base_name):
    """Busca un CSV por nombre exacto primero, luego variantes."""
    exact = os.path.join(data_dir, f"{base_name}.csv")
    if os.path.isfile(exact):
        return exact

    pattern = os.path.join(data_dir, f"{base_name}*.csv")
    candidates = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    return candidates[0] if candidates else None


def _load_csv(data_dir, filename):
    """Lee un CSV desde data_dir con encoding latin-1."""
    path = _find_csv(data_dir, filename)
    if not path:
        return pd.DataFrame()
    return pd.read_csv(path, encoding="latin-1")


# ── Filtrado de dreserva ─────────────────────────────────────────────────

def _load_filtered_dreserva(data_dir):
    """Carga dreserva.csv con los filtros de comisiones pendientes."""
    df = _load_csv(data_dir, "dreserva")
    if df.empty:
        return df

    mask = (
        (df["comision_pendiente"] == 1)
        & (df["monto_comision"] != 0)
        & (df["fecha_pago"].astype(str).str.strip() == "0000-00-00")
        & (df["servicio_cancelado"] == 0)
    )
    return df[mask].copy()


# ── Lookups ──────────────────────────────────────────────────────────────

def _load_proveedor_df(data_dir):
    """Carga proveedor.csv completo como DataFrame."""
    return _load_csv(data_dir, "proveedor")


def _build_proveedor_lookup(prov_df):
    """Crea {clave: {nombre, email, ciudad}} desde proveedor.csv."""
    lookup = {}
    if prov_df.empty:
        return lookup
    for _, r in prov_df.iterrows():
        clave = r.get("clave")
        if pd.notna(clave):
            lookup[int(clave)] = {
                "nombre": str(r.get("nombre", "")).strip() if pd.notna(r.get("nombre")) else "",
                "email": str(r.get("correo_e_contacto", "")).strip() if pd.notna(r.get("correo_e_contacto")) else "",
                "ciudad": str(r.get("ciudad", "")).strip() if pd.notna(r.get("ciudad")) else "",
            }
    return lookup


def _build_vendedor_lookup(data_dir):
    """Crea {folio: vendedor} desde reserva.csv."""
    df = _load_csv(data_dir, "reserva")
    if df.empty:
        return {}
    lookup = {}
    for _, r in df.iterrows():
        folio = r.get("folio")
        vendedor = r.get("vendedor", "")
        if pd.notna(folio):
            lookup[int(folio)] = str(vendedor).strip()
    return lookup


# ── Tabla FX ─────────────────────────────────────────────────────────────

def _build_fx_rows(fx):
    """Construye filas FX: [(DIVISA, Fx_EUR, Fx_USD), ...]."""
    rows = []
    for cur in _FX_CURRENCIES:
        fx_eur = fx.get(cur, {}).get("EUR", FALLBACK_FX.get(cur, {}).get("EUR", 1))
        fx_usd = fx.get(cur, {}).get("USD", FALLBACK_FX.get(cur, {}).get("USD", 1))
        rows.append((cur, fx_eur, fx_usd))
    return rows


def _get_fx_eur(moneda, fx):
    """Obtiene tasa moneda->EUR."""
    rate = fx.get(moneda, {}).get("EUR")
    if rate:
        return rate
    return FALLBACK_FX.get(moneda, {}).get("EUR", 1)


def _get_fx_usd(moneda, fx):
    """Obtiene tasa moneda->USD."""
    rate = fx.get(moneda, {}).get("USD")
    if rate:
        return rate
    return FALLBACK_FX.get(moneda, {}).get("USD", 1)


# ── Helpers de fecha ─────────────────────────────────────────────────────

def _parse_date(val):
    """Parsea un valor de fecha, retorna datetime o None."""
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s in ("", "0000-00-00"):
        return None
    try:
        return pd.to_datetime(s)
    except Exception:
        return None


# ── Escritura de hojas de datos ──────────────────────────────────────────
#
# Data LLC layout:
#   A: folio  B: proveedor  C: inicio_estancia  D: fin_estancia
#   E: moneda  F: monto_comision
#   G: =F*XLOOKUP(E,Z:Z,AA:AA)         Comision en EUR
#   H: =D+45                            45 dias despues fin
#   I: =TEXT(H,"mmmm")                  mes 45 dias
#   J: =MONTH(H)                        Numero Mes
#   K: =YEAR(H)                         ano 45 dias
#   L: =XLOOKUP(B,Q:Q,R:R)             Nombre Proveedor
#   M: =XLOOKUP(A,V:V,W:W)             Nombre vendedor
#   N: =F*XLOOKUP(E,Z:Z,AB:AB)         Comision en USD  ← NUEVO
#   Q-R: clave, nombre (proveedor lookup)
#   V-W: folio, vendedor (vendedor lookup)
#   Z-AB: DIVISA, Fx EUR, Fx USD
#
# Data SL layout:
#   A-F: mismo que LLC
#   G: =F*XLOOKUP(E,X:X,Y:Y)           Comision en EUR
#   H: =XLOOKUP(B,S:S,T:T)             Nombre Provedor
#   I: =D+45                            45 dias despues fin
#   J: =TEXT(I,"mmmm")                  mes 45 dias
#   K: =MONTH(I)                        Numero mes
#   L: =YEAR(I)                         ano 45 dias
#   M: =XLOOKUP(A,P:P,Q:Q,"no")        Vendedor
#   N: =F*XLOOKUP(E,X:X,Z:Z)           Comision en USD  ← NUEVO
#   P-Q: folio, vendedor (vendedor lookup)
#   S-T: codigo proveedor, nombre (proveedor lookup)
#   X-Z: DIVISA, Fx EUR, Fx USD

def _clear_data_rows(ws):
    """Limpia todos los datos debajo de la fila 1 (preserva headers)."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def _write_data_llc(ws, dreserva_df, prov_df, vendedor_lookup, fx):
    """Escribe datos en la hoja Data LLC."""
    _clear_data_rows(ws)
    if dreserva_df.empty:
        return

    # Columnas A-F (datos directos) + G-M (formulas)
    for i, (_, row) in enumerate(dreserva_df.iterrows(), start=2):
        ws.cell(i, 1, row.get("folio"))        # A: folio
        ws.cell(i, 2, row.get("proveedor"))    # B: proveedor

        dt_inicio = _parse_date(row.get("inicio_estancia"))
        if dt_inicio:
            ws.cell(i, 3, dt_inicio)
            ws.cell(i, 3).number_format = "DD/MM/YYYY"

        dt_fin = _parse_date(row.get("fin_estancia"))
        if dt_fin:
            ws.cell(i, 4, dt_fin)
            ws.cell(i, 4).number_format = "DD/MM/YYYY"

        ws.cell(i, 5, str(row.get("moneda", "")).strip())   # E: moneda
        ws.cell(i, 6, row.get("monto_comision"))             # F: monto_comision

        # Formulas
        ws.cell(i, 7).value = f"=F{i}*_xlfn.XLOOKUP(E{i},Z:Z,AA:AA)"       # G
        ws.cell(i, 8).value = f"=D{i}+45"                                    # H
        ws.cell(i, 8).number_format = "DD/MM/YYYY"
        ws.cell(i, 9).value = f'=TEXT(H{i},"mmmm")'                          # I
        ws.cell(i, 10).value = f"=MONTH(H{i})"                               # J
        ws.cell(i, 11).value = f"=YEAR(H{i})"                                # K
        ws.cell(i, 12).value = f"=_xlfn.XLOOKUP(B{i},Q:Q,R:R)"              # L
        ws.cell(i, 13).value = f"=_xlfn.XLOOKUP(A{i},V:V,W:W)"              # M
        ws.cell(i, 14).value = f"=F{i}*_xlfn.XLOOKUP(E{i},Z:Z,AB:AB)"      # N: USD

    # Header columna N
    ws.cell(1, 14).value = "Comision en USD"

    # Proveedor lookup Q-R
    if not prov_df.empty:
        for j, (_, r) in enumerate(prov_df.iterrows(), start=2):
            ws.cell(j, 17, r.get("clave"))     # Q: clave
            ws.cell(j, 18, r.get("nombre"))    # R: nombre

    # Vendedor lookup V-W
    for j, (folio, vend) in enumerate(sorted(vendedor_lookup.items()), start=2):
        ws.cell(j, 22, folio)      # V: folio
        ws.cell(j, 23, vend)       # W: vendedor

    # FX table Z-AB
    for j, (divisa, fx_eur, fx_usd) in enumerate(_build_fx_rows(fx), start=2):
        ws.cell(j, 26, divisa)     # Z: DIVISA
        ws.cell(j, 27, fx_eur)     # AA: Fx EUR
        ws.cell(j, 28, fx_usd)     # AB: Fx USD


def _write_data_sl(ws, dreserva_df, prov_df, vendedor_lookup, fx):
    """Escribe datos en la hoja Data SL."""
    _clear_data_rows(ws)
    if dreserva_df.empty:
        return

    # Columnas A-F (datos directos) + G-M (formulas - layout SL)
    for i, (_, row) in enumerate(dreserva_df.iterrows(), start=2):
        ws.cell(i, 1, row.get("folio"))        # A: folio
        ws.cell(i, 2, row.get("proveedor"))    # B: proveedor

        dt_inicio = _parse_date(row.get("inicio_estancia"))
        if dt_inicio:
            ws.cell(i, 3, dt_inicio)
            ws.cell(i, 3).number_format = "DD/MM/YYYY"

        dt_fin = _parse_date(row.get("fin_estancia"))
        if dt_fin:
            ws.cell(i, 4, dt_fin)
            ws.cell(i, 4).number_format = "DD/MM/YYYY"

        ws.cell(i, 5, str(row.get("moneda", "")).strip())   # E: moneda
        ws.cell(i, 6, row.get("monto_comision"))             # F: monto_comision

        # Formulas (SL tiene layout diferente: nombre prov antes de fecha)
        ws.cell(i, 7).value = f"=F{i}*_xlfn.XLOOKUP(E{i},X:X,Y:Y)"         # G: EUR
        ws.cell(i, 8).value = f"=_xlfn.XLOOKUP(B{i},S:S,T:T)"              # H: Nombre Prov
        ws.cell(i, 9).value = f"=D{i}+45"                                    # I: 45 dias
        ws.cell(i, 9).number_format = "DD/MM/YYYY"
        ws.cell(i, 10).value = f'=TEXT(I{i},"mmmm")'                         # J: mes
        ws.cell(i, 11).value = f"=MONTH(I{i})"                               # K: num mes
        ws.cell(i, 12).value = f"=YEAR(I{i})"                                # L: ano
        ws.cell(i, 13).value = f'=_xlfn.XLOOKUP(A{i},P:P,Q:Q,"no")'         # M: Vendedor
        ws.cell(i, 14).value = f"=F{i}*_xlfn.XLOOKUP(E{i},X:X,Z:Z)"        # N: USD

    # Header columna N
    ws.cell(1, 14).value = "Comision en USD"

    # Vendedor lookup P-Q
    for j, (folio, vend) in enumerate(sorted(vendedor_lookup.items()), start=2):
        ws.cell(j, 16, folio)      # P: folio
        ws.cell(j, 17, vend)       # Q: vendedor

    # Proveedor lookup S-T
    if not prov_df.empty:
        for j, (_, r) in enumerate(prov_df.iterrows(), start=2):
            ws.cell(j, 19, r.get("clave"))     # S: codigo proveedor
            ws.cell(j, 20, r.get("nombre"))    # T: nombre

    # FX table X-Z
    for j, (divisa, fx_eur, fx_usd) in enumerate(_build_fx_rows(fx), start=2):
        ws.cell(j, 24, divisa)     # X: DIVISA
        ws.cell(j, 25, fx_eur)     # Y: Fx EUR
        ws.cell(j, 26, fx_usd)     # Z: Fx USD


# ── Hojas Desglose (una por entidad) ─────────────────────────────────────

_DESGLOSE_HEADERS = [
    "Proveedor", "Nombre Proveedor", "Email Contacto", "Ciudad",
    "Folio", "Fin Estancia", "Moneda", "Monto Comision", "Comision EUR",
    "Comision USD", "Fecha Limite (45d)", "Mes", "Ano", "Vendedor",
]


def _build_desglose_rows(dreserva_df, prov_lookup, vendedor_lookup, fx):
    """Construye filas del desglose para una entidad, ordenadas por proveedor."""
    rows = []
    for _, row in dreserva_df.iterrows():
        folio = row.get("folio")
        prov_code = row.get("proveedor")
        moneda = str(row.get("moneda", "")).strip()
        monto = row.get("monto_comision", 0)

        prov_info = prov_lookup.get(int(prov_code), {}) if pd.notna(prov_code) else {}
        vendedor = vendedor_lookup.get(int(folio), "") if pd.notna(folio) else ""

        dt_fin = _parse_date(row.get("fin_estancia"))
        fecha_limite = (dt_fin + timedelta(days=45)) if dt_fin else None
        mes = fecha_limite.month if fecha_limite else None
        ano = fecha_limite.year if fecha_limite else None

        fx_eur = _get_fx_eur(moneda, fx)
        fx_usd = _get_fx_usd(moneda, fx)
        comision_eur = round(monto * fx_eur, 2) if monto else 0
        comision_usd = round(monto * fx_usd, 2) if monto else 0

        rows.append({
            "proveedor": prov_code,
            "nombre_prov": prov_info.get("nombre", ""),
            "email": prov_info.get("email", ""),
            "ciudad": prov_info.get("ciudad", ""),
            "folio": folio,
            "fin": dt_fin,
            "moneda": moneda,
            "monto": monto,
            "comision_eur": comision_eur,
            "comision_usd": comision_usd,
            "fecha_limite": fecha_limite,
            "mes": mes,
            "ano": ano,
            "vendedor": vendedor,
        })

    # Ordenar por nombre proveedor, luego folio
    rows.sort(key=lambda r: (r.get("nombre_prov", ""), r.get("folio") or 0))
    return rows


def _write_desglose_sheet(wb, sheet_name, entity_label, desglose_rows):
    """Crea una hoja Desglose para una entidad."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    title_font = Font(bold=True, size=13, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    n_cols = len(_DESGLOSE_HEADERS)

    # Titulo
    ws.cell(1, 1).value = f"Comisiones Pendientes - {entity_label} ({len(desglose_rows)} registros)"
    ws.cell(1, 1).font = title_font
    for c in range(1, n_cols + 1):
        ws.cell(1, c).fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Headers
    for col_idx, header in enumerate(_DESGLOSE_HEADERS, start=1):
        cell = ws.cell(2, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    if not desglose_rows:
        ws.cell(3, 1).value = "Sin registros"
        return

    date_fmt = "DD/MM/YYYY"
    for i, row in enumerate(desglose_rows, start=3):
        ws.cell(i, 1, row["proveedor"])
        ws.cell(i, 2, row["nombre_prov"])
        ws.cell(i, 3, row["email"])
        ws.cell(i, 4, row["ciudad"])
        ws.cell(i, 5, row["folio"])

        if row["fin"]:
            ws.cell(i, 6, row["fin"])
            ws.cell(i, 6).number_format = date_fmt

        ws.cell(i, 7, row["moneda"])
        ws.cell(i, 8, row["monto"])
        ws.cell(i, 8).number_format = "#,##0.00"
        ws.cell(i, 9, row["comision_eur"])
        ws.cell(i, 9).number_format = "#,##0.00"
        ws.cell(i, 10, row["comision_usd"])
        ws.cell(i, 10).number_format = "#,##0.00"

        if row["fecha_limite"]:
            ws.cell(i, 11, row["fecha_limite"])
            ws.cell(i, 11).number_format = date_fmt
        if row["mes"]:
            ws.cell(i, 12, MONTH_NAMES_ES.get(row["mes"], str(row["mes"])))
        if row["ano"]:
            ws.cell(i, 13, row["ano"])

        ws.cell(i, 14, row["vendedor"])

        for col in range(1, n_cols + 1):
            ws.cell(i, col).border = thin_border

    # Anchos de columna
    widths = [10, 25, 30, 15, 10, 14, 8, 14, 14, 14, 14, 12, 8, 15]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ── Limpieza y pivot refresh ─────────────────────────────────────────────

def _cleanup_sheets(wb):
    """Elimina hojas innecesarias del template (Detail1, Com Pend SL old, etc.)."""
    to_delete = [sn for sn in wb.sheetnames if sn not in _KEEP_SHEETS]
    for sn in to_delete:
        del wb[sn]


def _style_pivot_sheets(wb):
    """Mantiene el estilo original de la plantilla en los pivots."""
    for ws_name in ["Com Pend LLC", "Com Pend SL"]:
        if ws_name not in wb.sheetnames:
            continue
        ws = wb[ws_name]

        for pt in getattr(ws, '_pivots', []):
            pt.pivotTableStyleInfo.name = "PivotStyleLight16"
            pt.pivotTableStyleInfo.showRowHeaders = True
            pt.pivotTableStyleInfo.showColHeaders = True
            pt.pivotTableStyleInfo.showRowStripes = False
            pt.pivotTableStyleInfo.showColStripes = False
            pt.pivotTableStyleInfo.showLastColumn = True

        print(f"  {ws_name}: estilo PivotStyleLight16 (template)")


def _update_pivot_source_range(wb, data_row_counts):
    """Actualiza el rango fuente de cada pivot para cubrir todos los datos.

    Args:
        wb: Workbook abierto.
        data_row_counts: {"Data LLC": n_rows, "Data SL": n_rows}
    """
    for ws in wb.worksheets:
        for pt in getattr(ws, '_pivots', []):
            cache = pt.cache
            src = cache.cacheSource
            if not src or not src.worksheetSource:
                continue

            sheet_name = src.worksheetSource.sheet
            n_rows = data_row_counts.get(sheet_name)
            if n_rows is None:
                continue

            # El rango debe cubrir header (fila 1) + todas las filas de datos
            # Columna N = Comision en USD (nueva)
            last_row = max(n_rows + 1, 2)  # minimo 2 (header + 1 fila)
            new_ref = f"A1:N{last_row}"
            src.worksheetSource.ref = new_ref
            print(f"  Pivot '{ws.title}' → rango actualizado: {new_ref}")


def _force_pivot_refresh(wb):
    """Fuerza el recalculo de pivots y formulas al abrir, con estilo distintivo."""
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True

    seen = set()
    for ws in wb.worksheets:
        for pt in getattr(ws, '_pivots', []):
            cache = pt.cache
            cid = id(cache)
            if cid in seen:
                continue
            seen.add(cid)

            cache.refreshOnLoad = True
            cache.recordCount = 0

            try:
                if cache.records is not None:
                    cache.records.r = []
            except (AttributeError, TypeError):
                pass



# ── Funcion principal ────────────────────────────────────────────────────

def generate_comisiones_report(entities, output_dir, fx, year=None, month=None):
    """
    Genera el Reporte de Comisiones Pendientes Proveedores.

    A diferencia de otros reportes, este genera UN solo archivo con ambas
    entidades (Data SL + Data LLC) en hojas separadas.

    Args:
        entities: Dict de entidades {key: {company, label, data_dir, ...}}.
        output_dir: Directorio de salida.
        fx: Dict de tipos de cambio {currency: {"EUR": rate, "USD": rate}}.
        year: Ano del reporte (default: ano actual).
        month: Mes del reporte (default: mes actual).

    Returns:
        Ruta del archivo generado o None si falla.
    """
    today = datetime.now()
    year = year or today.year
    month = month or today.month
    month_name = MONTH_NAMES_ES.get(month, str(month)).capitalize()

    # Verificar template
    if not os.path.isfile(COM_PEND_PROV_TEMPLATE_PATH):
        print(f"  ERROR: Template no encontrado: {COM_PEND_PROV_TEMPLATE_PATH}")
        return None

    # Copiar template
    filename = f"Comisiones_Pendientes_Prov_{month_name}_{year}.xlsx"
    filepath = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)
    shutil.copy2(COM_PEND_PROV_TEMPLATE_PATH, filepath)

    wb = load_workbook(filepath)

    # Limpiar hojas innecesarias
    _cleanup_sheets(wb)
    print(f"  Hojas: {wb.sheetnames}")

    # Mapping: data sheet name → number of data rows written
    data_row_counts = {}

    for key, cfg in entities.items():
        company = cfg["company"]
        data_dir = cfg["data_dir"]
        entity_label = _ENTITY_LABELS.get(company, cfg["label"])

        if not os.path.exists(data_dir):
            print(f"  {entity_label}: directorio no encontrado ({data_dir})")
            continue

        # Cargar datos
        dreserva_df = _load_filtered_dreserva(data_dir)
        prov_df = _load_proveedor_df(data_dir)
        prov_lookup = _build_proveedor_lookup(prov_df)
        vendedor_lookup = _build_vendedor_lookup(data_dir)

        n_rows = len(dreserva_df)
        print(f"  {entity_label}: {n_rows} comisiones pendientes")

        # Escribir hoja de datos
        if company == "LLC" and "Data LLC" in wb.sheetnames:
            _write_data_llc(wb["Data LLC"], dreserva_df, prov_df, vendedor_lookup, fx)
            data_row_counts["Data LLC"] = n_rows
        elif company == "SL" and "Data SL" in wb.sheetnames:
            _write_data_sl(wb["Data SL"], dreserva_df, prov_df, vendedor_lookup, fx)
            data_row_counts["Data SL"] = n_rows

        # Desglose por entidad
        desglose_rows = _build_desglose_rows(dreserva_df, prov_lookup, vendedor_lookup, fx)
        sheet_name = f"Desglose {company}"
        _write_desglose_sheet(wb, sheet_name, entity_label, desglose_rows)
        print(f"  {sheet_name}: {len(desglose_rows)} registros")

    # Actualizar rangos de pivot tables al numero real de filas
    _update_pivot_source_range(wb, data_row_counts)

    # FX Rates (tasas diarias del mes)
    month_label, n_days = write_fx_sheet(wb)
    print(f"  FX Rates: {n_days} dias de {month_label}")

    # Force pivot refresh
    _force_pivot_refresh(wb)
    print(f"  Force refresh configurado")

    # Colores personalizados en pivots
    _style_pivot_sheets(wb)

    wb.save(filepath)
    wb.close()

    return filepath
