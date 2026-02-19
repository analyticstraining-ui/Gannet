"""
Genera el Reporte Mensual Consolidado de Travel Advisors.

Flujo:
  1. Copiar template
  2. Escribir datos enriquecidos en hoja DATA NEW
  3. Actualizar filtros de las tablas dinamicas (mes/ano) para el periodo actual
  4. Configurar refreshOnLoad para que los pivots se refresquen al abrir en Excel
"""

import os
import csv
import shutil
from collections import OrderedDict
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties

from config import BASE_DIR, MONTH_NAMES_ES, TA_TEMPLATE_PATH
from src.fx_rates import get_current_month_daily_rates

# ── Estilos ──────────────────────────────────────────────────────────────
FONT = Font(name="Aptos Narrow", size=11)
FONT_BOLD = Font(name="Aptos Narrow", size=11, bold=True)
FONT_HDR = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")

THIN = Side(style="thin", color="BFBFBF")
MEDIUM = Side(style="medium", color="808080")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_BORDER = Border(left=THIN, right=THIN, top=MEDIUM, bottom=MEDIUM)

DATA_NEW_HDR = PatternFill("solid", fgColor="4472C4")
HDR_FILL = PatternFill("solid", fgColor="548235")
GREEN_FILL = PatternFill("solid", fgColor="92D050")

PCT_FMT = '0.0%'
DATE_FMT = 'DD/MM/YYYY'
NUM_FMT = '#,##0.00'
REPORT_NUM_FMT = '[$-409]#,##0'

# FX lookup table position: cols AO-AQ (41-43)
FX_COL_START = 41  # AO

# Definicion de columnas DATA NEW
DATA_NEW_COLS = [
    ("A", "Compania"), ("B", "folio"), ("C", "cerrada"),
    ("D", "fecha"), ("E", "fecha_inicio"), ("F", "fecha_fin"),
    ("G", "vendedor"), ("H", "Semana"), ("I", "usuarios_invitados"),
    ("J", "total_cliente"), ("K", "moneda"),
    ("L", "Total Venta EUR"), ("M", "Total Venta USD"),
    ("N", "Rentabilidad"), ("O", "Rentabilidad en EUR"),
    ("P", "Rentabilidad en USD"), ("Q", "% Rentabilidad"),
    ("R", "Mes"), ("S", "Ano"), ("T", "Mes Inicio"), ("U", "Ano Inicio"),
    ("V", "Fecha 45 dias fin"), ("W", "mes 45 dias fin"),
    ("X", "Ano 45 dias fin"),
    ("Y_oficina", "Oficina"), ("Z_linea", "Linea de Negocio"),
    ("AA_com", "Comisionamiento"),
    ("AB_renta_com", "Renta after COM"),
    ("AC_renta_com_usd", "Renta after COM USD"),
    ("AD_fecha_inc", "Fecha Incorporacion"),
]

DATA_NEW_WIDTHS = [10, 8, 7, 12, 12, 12, 14, 8, 10, 14, 8,
                   16, 16, 14, 16, 16, 14, 6, 6, 10, 10,
                   16, 14, 10, 12, 16, 16, 16, 18, 16]


# ── Helpers ──────────────────────────────────────────────────────────────

def _cell(ws, row, col, val, font=FONT, fmt=None, fill=None, border=None,
          indent=0, alignment=None):
    """Escribe un valor con formato completo."""
    cell = ws.cell(row, col, val)
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if indent:
        cell.alignment = Alignment(indent=indent, horizontal="left")
    elif alignment:
        cell.alignment = alignment
    return cell


# ── Plantilla Corsario ───────────────────────────────────────────────────

def load_plantilla_corsario():
    """Carga la Plantilla Corsario desde el CSV en templates/."""
    path = os.path.join(BASE_DIR, "templates", "plantilla_corsario.csv")
    plantilla = []
    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            com = row["comisionamiento"]
            try:
                com = float(com) if com else 0
            except ValueError:
                com = 0
            fecha = None
            if row["fecha_inicio"]:
                try:
                    fecha = datetime.strptime(row["fecha_inicio"], "%Y-%m-%d")
                except ValueError:
                    pass
            plantilla.append({
                "usuario": row["usuario_corsario"].strip(),
                "linea_negocio": row["linea_negocio"].strip(),
                "comisionamiento": com,
                "oficina": row["pais_trabajo"].strip(),
                "fecha_inicio": fecha,
            })
    return plantilla


def _build_lookup(plantilla):
    """Crea dict {vendedor: {oficina, linea_negocio, comisionamiento, fecha_inicio}}."""
    return {p["usuario"]: p for p in plantilla}


# ── Enriquecer data_rows ─────────────────────────────────────────────────

def enrich_data_rows(data_rows, plantilla, fx):
    """Agrega columnas Y-AD y recalcula L, M, O, P, Q, AB, AC con fx_latest.

    Las columnas de conversion FX se calculan aqui con los rates actuales
    (no historicos) para que los pivots tengan valores reales.
    """
    lookup = _build_lookup(plantilla)
    enriched = []
    for r in data_rows:
        row = dict(r)  # copia
        vendedor = str(row.get("G") or "").strip()
        info = lookup.get(vendedor, {})

        # Plantilla Corsario
        row["Y_oficina"] = info.get("oficina", "")
        row["Z_linea"] = info.get("linea_negocio", "")
        row["AA_com"] = info.get("comisionamiento", 0)
        row["AD_fecha_inc"] = info.get("fecha_inicio")

        # Conversion FX con tasas actuales
        moneda = str(row.get("K") or "EUR").strip()
        total = float(row.get("J") or 0)
        rent = float(row.get("N") or 0)
        fx_rates = fx.get(moneda, fx.get("EUR", {"EUR": 1.0, "USD": 1.0}))
        fx_eur = fx_rates["EUR"]
        fx_usd = fx_rates["USD"]

        row["L"] = round(total * fx_eur, 2)
        row["M"] = round(total * fx_usd, 2)
        row["O"] = round(rent * fx_eur, 2)
        row["P"] = round(rent * fx_usd, 2)
        row["Q"] = round(rent / total, 6) if total != 0 else 0

        com = float(row["AA_com"] or 0)
        row["AB_renta_com"] = round(row["O"] * (1 - com), 2)
        row["AC_renta_com_usd"] = round(row["P"] * (1 - com), 2)

        enriched.append(row)
    return enriched


# ── Plantilla Corsario en hoja ───────────────────────────────────────────

def _write_plantilla_block(ws, plantilla, col_start=35, row_start=1):
    """Escribe la plantilla corsario en cols AI-AM (35-39) desde row_start."""
    headers = ["usuario Corsario", "Linea de Negocio", "Comisionamiento",
               "Pais de trabajo", "Fecha Inicio"]
    for i, h in enumerate(headers):
        _cell(ws, row_start, col_start + i, h,
              font=FONT_HDR, fill=HDR_FILL, border=HDR_BORDER)

    for j, p in enumerate(plantilla):
        r = row_start + 1 + j
        _cell(ws, r, col_start, p["usuario"], border=THIN_BORDER)
        _cell(ws, r, col_start + 1, p["linea_negocio"], border=THIN_BORDER)
        _cell(ws, r, col_start + 2, p["comisionamiento"], fmt=PCT_FMT, border=THIN_BORDER)
        _cell(ws, r, col_start + 3, p["oficina"], border=THIN_BORDER)
        if p["fecha_inicio"]:
            _cell(ws, r, col_start + 4, p["fecha_inicio"], fmt=DATE_FMT, border=THIN_BORDER)
        else:
            _cell(ws, r, col_start + 4, None, border=THIN_BORDER)

    ws.column_dimensions[get_column_letter(col_start)].width = 15
    ws.column_dimensions[get_column_letter(col_start + 1)].width = 15.5
    ws.column_dimensions[get_column_letter(col_start + 2)].width = 16.5
    ws.column_dimensions[get_column_letter(col_start + 3)].width = 13.5
    ws.column_dimensions[get_column_letter(col_start + 4)].width = 11.5


# ── FX lookup table ───────────────────────────────────────────────────

def _write_fx_table(ws, fx):
    """Escribe la tabla de tipos de cambio en cols AO-AQ (41-43)."""
    _cell(ws, 1, FX_COL_START, "Moneda",
          font=FONT_HDR, fill=HDR_FILL, border=HDR_BORDER)
    _cell(ws, 1, FX_COL_START + 1, "FX EUR",
          font=FONT_HDR, fill=HDR_FILL, border=HDR_BORDER)
    _cell(ws, 1, FX_COL_START + 2, "FX USD",
          font=FONT_HDR, fill=HDR_FILL, border=HDR_BORDER)

    currencies = ["EUR", "USD", "CHF", "GBP", "GPB", "JPY", "MXN"]
    for i, curr in enumerate(currencies):
        r = 2 + i
        _cell(ws, r, FX_COL_START, curr, border=THIN_BORDER)
        if curr in fx:
            _cell(ws, r, FX_COL_START + 1, fx[curr]["EUR"],
                  fmt='0.000000', border=THIN_BORDER)
            _cell(ws, r, FX_COL_START + 2, fx[curr]["USD"],
                  fmt='0.000000', border=THIN_BORDER)

    ws.column_dimensions[get_column_letter(FX_COL_START)].width = 10
    ws.column_dimensions[get_column_letter(FX_COL_START + 1)].width = 12
    ws.column_dimensions[get_column_letter(FX_COL_START + 2)].width = 12


# ── Escribir DATA NEW ───────────────────────────────────────────────────

def _write_data_new(ws, enriched, plantilla, fx):
    """Escribe datos enriquecidos en DATA NEW con valores pre-calculados."""

    # Limpiar datos existentes (preservar estructura)
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(DATA_NEW_COLS) + 1):
            ws.cell(row, col).value = None

    # Verificar/escribir headers
    for ci, (_, header) in enumerate(DATA_NEW_COLS, 1):
        _cell(ws, 1, ci, header, font=FONT_HDR, fill=DATA_NEW_HDR, border=HDR_BORDER)

    # Escribir datos (todos como valores para compatibilidad con pivots)
    for ri, row_data in enumerate(enriched, 2):
        for ci, (key, _) in enumerate(DATA_NEW_COLS, 1):
            val = row_data.get(key)
            if val is not None:
                fmt = None
                if key in ("D", "E", "F", "V", "AD_fecha_inc"):
                    if isinstance(val, datetime):
                        fmt = DATE_FMT
                elif key in ("J", "L", "M", "N", "O", "P", "AB_renta_com", "AC_renta_com_usd"):
                    fmt = NUM_FMT
                elif key in ("Q", "AA_com"):
                    fmt = PCT_FMT
                _cell(ws, ri, ci, val, fmt=fmt, border=THIN_BORDER)
            else:
                _cell(ws, ri, ci, None, border=THIN_BORDER)

    # Anchos
    for i, w in enumerate(DATA_NEW_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Plantilla Corsario en cols AI-AM (35-39)
    _write_plantilla_block(ws, plantilla, col_start=35, row_start=1)

    # FX lookup table en cols AO-AQ (41-43) — referencia de tasas usadas
    _write_fx_table(ws, fx)


# ── Report helpers ─────────────────────────────────────────────────────

def _build_fecha_inc_lookup(enriched):
    """Build {vendor_name: fecha_incorporacion} from enriched data."""
    lookup = {}
    for r in enriched:
        v = str(r.get("G") or "").strip()
        if v and v not in lookup:
            fi = r.get("AD_fecha_inc")
            if fi:
                lookup[v] = fi
    return lookup


def _build_hierarchy(rows, value_keys):
    """Group rows by Oficina -> LN -> Vendedor, summing value_keys.

    Returns (hierarchy OrderedDict, grand_totals dict).
    """
    tree = {}
    for r in rows:
        oficina = r.get("Y_oficina") or "(en blanco)"
        ln = r.get("Z_linea") or "(en blanco)"
        vendor = str(r.get("G") or "").strip()
        if not vendor:
            continue
        tree.setdefault(oficina, {}).setdefault(ln, {}).setdefault(
            vendor, {k: 0.0 for k in value_keys})
        for k in value_keys:
            val = r.get(k)
            if val is not None:
                tree[oficina][ln][vendor][k] += float(val)

    hierarchy = OrderedDict()
    grand = {k: 0.0 for k in value_keys}

    for oficina in sorted(tree):
        ofi_totals = {k: 0.0 for k in value_keys}
        ofi_lineas = OrderedDict()

        for ln in sorted(tree[oficina]):
            ln_totals = {k: 0.0 for k in value_keys}
            ln_vendors = OrderedDict()

            for vendor in sorted(tree[oficina][ln]):
                vals = tree[oficina][ln][vendor]
                ln_vendors[vendor] = vals
                for k in value_keys:
                    ln_totals[k] += vals[k]

            ofi_lineas[ln] = {"totals": ln_totals, "vendors": ln_vendors}
            for k in value_keys:
                ofi_totals[k] += ln_totals[k]

        hierarchy[oficina] = {"totals": ofi_totals, "lineas": ofi_lineas}
        for k in value_keys:
            grand[k] += ofi_totals[k]

    return hierarchy, grand


def _write_hierarchy_block(ws, hierarchy, grand, value_keys, start_row,
                           label_col, val_start_col, num_fmt=REPORT_NUM_FMT,
                           fecha_inc_col=None, fecha_inc_lookup=None):
    """Write hierarchy (Oficina->LN->Vendor) block with indentation.

    Returns next_row (row after Total general).
    """
    r = start_row

    for oficina, ofi_data in hierarchy.items():
        # Office row (indent=0, bold)
        _cell(ws, r, label_col, oficina, font=FONT_BOLD, indent=0)
        for i, k in enumerate(value_keys):
            val = round(ofi_data["totals"][k], 2)
            _cell(ws, r, val_start_col + i, val if val else None,
                  font=FONT_BOLD, fmt=num_fmt)
        r += 1

        for ln, ln_data in ofi_data["lineas"].items():
            # LN row (indent=1, bold)
            _cell(ws, r, label_col, ln, font=FONT_BOLD, indent=1)
            for i, k in enumerate(value_keys):
                val = round(ln_data["totals"][k], 2)
                _cell(ws, r, val_start_col + i, val if val else None,
                      font=FONT_BOLD, fmt=num_fmt)
            r += 1

            for vendor, vals in ln_data["vendors"].items():
                # Vendor row (indent=2)
                _cell(ws, r, label_col, vendor, indent=2)
                for i, k in enumerate(value_keys):
                    val = round(vals[k], 2)
                    _cell(ws, r, val_start_col + i, val if val else None,
                          fmt=num_fmt)
                if fecha_inc_col and fecha_inc_lookup:
                    fi = fecha_inc_lookup.get(vendor)
                    if fi:
                        _cell(ws, r, fecha_inc_col, fi, fmt=DATE_FMT)
                r += 1

    # Total general
    _cell(ws, r, label_col, "Total general", font=FONT_BOLD, indent=0)
    for i, k in enumerate(value_keys):
        val = round(grand[k], 2)
        _cell(ws, r, val_start_col + i, val if val else None,
              font=FONT_BOLD, fmt=num_fmt)
    r += 1

    return r


def _clear_report_sheet(ws):
    """Clear all cell values and formatting in a report sheet."""
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None
            cell.font = FONT
            cell.fill = PatternFill()
            cell.alignment = Alignment()
            cell.number_format = 'General'


# ── Report writing functions ─────────────────────────────────────────────

_REPORT_VALUE_KEYS = ["M", "AC_renta_com_usd", "P"]
_REPORT_HEADERS = ["Suma de Total Venta USD", "Suma de Renta after COM USD",
                   "Suma de Rentabilidad en USD"]


def _write_report_1_2(ws, enriched, year, month, month_name):
    """Write Reports 1 (current month) and 2 (YTD) side by side.

    Returns next_row after both reports.
    """
    # Filter data
    r1_data = [r for r in enriched if r.get("S") == year and r.get("R") == month]
    r2_data = [r for r in enriched if r.get("S") == year and r.get("R") <= month]

    # Build hierarchies
    h1, g1 = _build_hierarchy(r1_data, _REPORT_VALUE_KEYS)
    h2, g2 = _build_hierarchy(r2_data, _REPORT_VALUE_KEYS)

    # ── Report 1 filter labels (cols A-D) ──
    _cell(ws, 1, 1, "Compania")
    _cell(ws, 1, 2, "(Todas)")
    _cell(ws, 2, 1, "Ano")
    _cell(ws, 2, 2, year)
    _cell(ws, 2, 3, "Report 1", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, 2, 4, f"Venta mes de {month_name}", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, 3, 1, "Mes")
    _cell(ws, 3, 2, month)

    # Report 1 column headers (row 5)
    _cell(ws, 5, 1, "Etiquetas de fila", font=FONT_BOLD)
    for i, h in enumerate(_REPORT_HEADERS):
        _cell(ws, 5, 2 + i, h, font=FONT_BOLD)

    # Report 1 data (row 6+)
    r1_end = _write_hierarchy_block(ws, h1, g1, _REPORT_VALUE_KEYS,
                                    start_row=6, label_col=1, val_start_col=2)

    # ── Report 2 filter labels (cols K-N) ──
    _cell(ws, 1, 11, "Compania")
    _cell(ws, 1, 12, "(Todas)")
    _cell(ws, 2, 11, "Ano")
    _cell(ws, 2, 12, year)
    _cell(ws, 2, 13, "Report 2", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, 2, 14, f"Venta acumulada {year}", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, 3, 11, "Mes")
    _cell(ws, 3, 12, month)

    # Report 2 column headers (row 5)
    _cell(ws, 5, 11, "Etiquetas de fila", font=FONT_BOLD)
    for i, h in enumerate(_REPORT_HEADERS):
        _cell(ws, 5, 12 + i, h, font=FONT_BOLD)

    # Report 2 data (row 6+)
    r2_end = _write_hierarchy_block(ws, h2, g2, _REPORT_VALUE_KEYS,
                                    start_row=6, label_col=11, val_start_col=12)

    return max(r1_end, r2_end)


def _write_report_3(ws, enriched, year, start_row):
    """Write Report 3 - Venta por Mes de Inicio.

    Filter: Ano Inicio == year (all bookings with trips starting in year).
    Group: Vendor x Mes Inicio.
    Values: sum(Total Venta USD).

    Returns next_row.
    """
    # Filter: Ano Inicio == year
    filtered = [r for r in enriched if r.get("U") == year]

    # Aggregate: vendor x mes_inicio -> sum(M)
    vendor_months = {}
    for r in filtered:
        vendor = str(r.get("G") or "").strip()
        if not vendor:
            continue
        mes = r.get("T")
        if mes is None:
            continue
        mes = int(mes)
        vendor_months.setdefault(vendor, {})
        vendor_months[vendor][mes] = vendor_months[vendor].get(mes, 0.0) + float(r.get("M") or 0)

    # Filter labels
    _cell(ws, start_row, 1, "Compania")
    _cell(ws, start_row, 2, "(Todas)")
    _cell(ws, start_row + 1, 1, "Ano Inicio")
    _cell(ws, start_row + 1, 2, year)
    _cell(ws, start_row + 2, 1, "Oficina")
    _cell(ws, start_row + 2, 2, "(Todas)")
    _cell(ws, start_row + 2, 3, "Report 3 ", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, start_row + 2, 4, "Venta Mes de Inicio", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, start_row + 3, 1, "Linea de Negocio")
    _cell(ws, start_row + 3, 2, "(Todas)")

    # Summary label
    _cell(ws, start_row + 5, 1, "Suma de Total Venta USD")
    _cell(ws, start_row + 5, 2, "Etiquetas de columna")

    # Headers: months 1-12 + Total general
    hdr_row = start_row + 6
    _cell(ws, hdr_row, 1, "Etiquetas de fila", font=FONT_BOLD)
    for m in range(1, 13):
        _cell(ws, hdr_row, 1 + m, m, font=FONT_BOLD)
    _cell(ws, hdr_row, 14, "Total general", font=FONT_BOLD)

    # Data rows
    r = hdr_row + 1
    month_totals = {m: 0.0 for m in range(1, 13)}
    grand_total = 0.0

    for vendor in sorted(vendor_months):
        _cell(ws, r, 1, vendor, indent=1)
        vendor_total = 0.0
        for m in range(1, 13):
            val = vendor_months[vendor].get(m)
            if val:
                _cell(ws, r, 1 + m, round(val, 2), fmt=REPORT_NUM_FMT)
                month_totals[m] += val
                vendor_total += val
        _cell(ws, r, 14, round(vendor_total, 2), fmt=REPORT_NUM_FMT)
        grand_total += vendor_total
        r += 1

    # Total general row
    _cell(ws, r, 1, "Total general", font=FONT_BOLD, indent=0)
    for m in range(1, 13):
        val = round(month_totals[m], 2)
        if val:
            _cell(ws, r, 1 + m, val, font=FONT_BOLD, fmt=REPORT_NUM_FMT)
    _cell(ws, r, 14, round(grand_total, 2), font=FONT_BOLD, fmt=REPORT_NUM_FMT)
    r += 1

    return r


def _write_report_4(ws, enriched, year, start_row, fecha_inc_lookup):
    """Write Report 4 - Venta por Mes with Fecha Incorporacion.

    Filter: Ano == year.
    Group: Vendor x Mes (calendar month of booking).
    Values: sum(Total Venta USD).
    Col A: Fecha Incorporacion.

    Returns next_row.
    """
    # Filter: Ano == year
    filtered = [r for r in enriched if r.get("S") == year]

    # Aggregate: vendor x mes -> sum(M)
    vendor_months = {}
    all_months = set()
    for r in filtered:
        vendor = str(r.get("G") or "").strip()
        if not vendor:
            continue
        mes = r.get("R")
        if mes is None:
            continue
        mes = int(mes)
        all_months.add(mes)
        vendor_months.setdefault(vendor, {})
        vendor_months[vendor][mes] = vendor_months[vendor].get(mes, 0.0) + float(r.get("M") or 0)

    months_sorted = sorted(all_months)
    n_months = len(months_sorted)

    # Filter labels (col B onwards)
    _cell(ws, start_row, 2, "Compania")
    _cell(ws, start_row, 3, "(Todas)")
    _cell(ws, start_row + 1, 2, "Ano")
    _cell(ws, start_row + 1, 3, year)
    _cell(ws, start_row + 1, 4, "Report 4", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, start_row + 1, 5, "Venta por Mes", font=FONT_BOLD, fill=GREEN_FILL)

    # Summary label
    _cell(ws, start_row + 3, 2, "Suma de Total Venta USD")
    _cell(ws, start_row + 3, 3, "Etiquetas de columna")

    # Headers
    hdr_row = start_row + 4
    _cell(ws, hdr_row, 2, "Etiquetas de fila", font=FONT_BOLD)
    for i, m in enumerate(months_sorted):
        _cell(ws, hdr_row, 3 + i, m, font=FONT_BOLD)
    total_col = 3 + n_months
    _cell(ws, hdr_row, total_col, "Total general", font=FONT_BOLD)

    # Data rows
    r = hdr_row + 1
    month_totals = {m: 0.0 for m in months_sorted}
    grand_total = 0.0

    for vendor in sorted(vendor_months):
        # Fecha incorporacion in col A
        fi = fecha_inc_lookup.get(vendor)
        if fi:
            _cell(ws, r, 1, fi, fmt=DATE_FMT)

        _cell(ws, r, 2, vendor, indent=1)
        vendor_total = 0.0
        for i, m in enumerate(months_sorted):
            val = vendor_months[vendor].get(m)
            if val:
                _cell(ws, r, 3 + i, round(val, 2), fmt=REPORT_NUM_FMT)
                month_totals[m] += val
                vendor_total += val
        _cell(ws, r, total_col, round(vendor_total, 2), fmt=REPORT_NUM_FMT)
        grand_total += vendor_total
        r += 1

    # Total general
    _cell(ws, r, 2, "Total general", font=FONT_BOLD, indent=0)
    for i, m in enumerate(months_sorted):
        val = round(month_totals[m], 2)
        if val:
            _cell(ws, r, 3 + i, val, font=FONT_BOLD, fmt=REPORT_NUM_FMT)
    _cell(ws, r, total_col, round(grand_total, 2), font=FONT_BOLD, fmt=REPORT_NUM_FMT)
    r += 1

    return r


def _write_ventas_ln(ws, enriched, year, month, fecha_inc_lookup):
    """Write the 'Ventas por Linea de Negocio' sheet.

    Filter: Ano == year AND Mes <= month (YTD).
    Hierarchy: Oficina -> LN -> Vendedor.
    """
    # Filter: YTD
    filtered = [r for r in enriched if r.get("S") == year and r.get("R") <= month]

    # Build hierarchy
    h, g = _build_hierarchy(filtered, _REPORT_VALUE_KEYS)

    # Filter labels
    _cell(ws, 1, 2, "Mes")
    _cell(ws, 1, 3, month)
    _cell(ws, 2, 2, "Compania")
    _cell(ws, 2, 3, "(Todas)")
    _cell(ws, 2, 4, f"Ventas por Linea de Negocio {year}",
          font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, 3, 2, "Ano")
    _cell(ws, 3, 3, year)
    _cell(ws, 4, 2, "Ano Inicio")
    _cell(ws, 4, 3, "(Todas)")

    # Column headers (row 6)
    _cell(ws, 6, 2, "Etiquetas de fila", font=FONT_BOLD)
    for i, h_text in enumerate(_REPORT_HEADERS):
        _cell(ws, 6, 3 + i, h_text, font=FONT_BOLD)

    # Hierarchy data with Fecha Incorporacion in col A
    _write_hierarchy_block(ws, h, g, _REPORT_VALUE_KEYS,
                           start_row=7, label_col=2, val_start_col=3,
                           fecha_inc_col=1, fecha_inc_lookup=fecha_inc_lookup)


# ── Pivot table filter updates ─────────────────────────────────────────

def _get_shared_items_mapping(cache, field_idx):
    """Parse shared items for a cache field.

    Returns {shared_item_index: numeric_or_string_value}.
    Skips missing (m) and error (e) entries.
    """
    cf = cache.cacheFields[field_idx]
    si_tree = cf.sharedItems.to_tree()
    mapping = {}
    for si_idx, elem in enumerate(si_tree):
        tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        v = elem.get('v')
        if tag == 'n' and v is not None:
            val = float(v)
            mapping[si_idx] = int(val) if val == int(val) else val
        elif tag == 's' and v is not None:
            mapping[si_idx] = v
    return mapping


def _get_pf_item_for_value(pt, field_idx, target_value):
    """Find the pivotField item index whose shared value matches target_value."""
    shared_map = _get_shared_items_mapping(pt.cache, field_idx)
    val_to_shared = {v: k for k, v in shared_map.items()}

    target_shared = val_to_shared.get(target_value)
    if target_shared is None:
        return None

    pf = pt.pivotFields[field_idx]
    for pf_idx, item in enumerate(pf.items):
        if item.x == target_shared:
            return pf_idx
    return None


def _set_pf_page_item(pt, field_idx, pf_item_idx):
    """Set a page field to select a specific item (single-select mode)."""
    if pt.pageFields:
        for pf in pt.pageFields:
            if pf.fld == field_idx:
                pf.item = pf_item_idx
                break


def _set_pf_multiselect(pt, field_idx, visible_values):
    """Set a page field to multi-select mode with specific values visible."""
    shared_map = _get_shared_items_mapping(pt.cache, field_idx)

    if pt.pageFields:
        for pf in pt.pageFields:
            if pf.fld == field_idx:
                pf.item = None
                break

    pivot_field = pt.pivotFields[field_idx]
    pivot_field.multipleItemSelectionAllowed = True

    for item in pivot_field.items:
        if item.x is not None and item.x in shared_map:
            if shared_map[item.x] in visible_values:
                item.h = None
            else:
                item.h = True


def _set_axis_items_visible(pt, field_idx, visible_values):
    """For a field on column/row axis, make specified values visible."""
    shared_map = _get_shared_items_mapping(pt.cache, field_idx)

    pivot_field = pt.pivotFields[field_idx]
    for item in pivot_field.items:
        if item.x is not None and item.x in shared_map:
            if shared_map[item.x] in visible_values:
                item.h = None
            else:
                item.h = True


# ── Layout adjustment for pivots ───────────────────────────────────────

# New layout positions to prevent pivot overlap after refresh.
# Reports 1&2 start at row 5 (cols A-D and K-N respectively).
# Report 3 moved from row 56 to row 120, Report 4 from row 102 to row 230.
_TD3_NEW_START = 120
_TD4_NEW_START = 230


def _write_plantilla_lookup(ws, plantilla, header_row, data_start_row):
    """Write plantilla corsario in cols AF-AJ for XLOOKUP formulas."""
    headers = ["usuario Corsario", "Linea de Negocio", "Comisionamiento",
               "País de trabajo", "Fecha Inicio"]
    for i, h in enumerate(headers):
        ws.cell(header_row, 32 + i, h)
    for j, p in enumerate(plantilla):
        r = data_start_row + j
        ws.cell(r, 32, p["usuario"])
        ws.cell(r, 33, p["linea_negocio"])
        ws.cell(r, 34, p["comisionamiento"])
        ws.cell(r, 35, p["oficina"])
        if p["fecha_inicio"]:
            c = ws.cell(r, 36, p["fecha_inicio"])
            c.number_format = DATE_FMT


def _prepare_report_sheet(ws, plantilla, month, year, month_name):
    """Clear cached pivot data, adjust layout, write labels and formulas."""
    # Clear all cells (cached pivot data + old labels)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Rewrite plantilla lookup in cols AF-AJ (header at row 4, data at row 5)
    _write_plantilla_lookup(ws, plantilla, header_row=4, data_start_row=5)

    # ── Decorative labels ──

    # Report 1 (cols A-D, pivot at A5)
    _cell(ws, 2, 3, "Report 1", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, 2, 4, f"Venta mes de {month_name}", font=FONT_BOLD, fill=GREEN_FILL)
    ws.cell(3, 2).value = month

    # Report 2 (cols K-N, pivot at K5)
    _cell(ws, 2, 13, "Report 2", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, 2, 14, f"Venta acumulada {year}", font=FONT_BOLD, fill=GREEN_FILL)
    ws.cell(3, 12).value = month

    # Report 3 (pivot at row _TD3_NEW_START)
    # Page fields: 4 fields → rows start-5 to start-2, gap at start-1
    r3_pf_start = _TD3_NEW_START - 5
    r3_label_row = r3_pf_start + 2  # 3rd page field row
    _cell(ws, r3_label_row, 3, "Report 3 ", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, r3_label_row, 4, "Venta Mes de Inicio", font=FONT_BOLD, fill=GREEN_FILL)

    # Report 4 (pivot at row _TD4_NEW_START)
    # Page fields: 2 fields → rows start-3 to start-2, gap at start-1
    r4_label_row = _TD4_NEW_START - 2  # 2nd page field row
    _cell(ws, r4_label_row, 4, "Report 4", font=FONT_BOLD, fill=GREEN_FILL)
    _cell(ws, r4_label_row, 5, "Venta por Mes", font=FONT_BOLD, fill=GREEN_FILL)

    # XLOOKUP formulas for Report 4 (Fecha Incorporacion in col A)
    # firstDataRow=2 → data starts at TD4_NEW_START + 2
    td4_data_start = _TD4_NEW_START + 2
    for r in range(td4_data_start, td4_data_start + 100):
        ws.cell(r, 1).value = f'=IFERROR(_xlfn.XLOOKUP(B{r},AF:AF,AJ:AJ),"")'

    # Adjust pivot location refs
    for pt in ws._pivots:
        if pt.name == "TablaDin\u00e1mica1":
            pt.location.ref = "A5:D110"
        elif pt.name == "TablaDin\u00e1mica2":
            pt.location.ref = "K5:N110"
        elif pt.name == "TablaDin\u00e1mica3":
            pt.location.ref = f"A{_TD3_NEW_START}:N{_TD3_NEW_START + 100}"
        elif pt.name == "TablaDin\u00e1mica4":
            pt.location.ref = f"B{_TD4_NEW_START}:Z{_TD4_NEW_START + 100}"


def _prepare_ventas_sheet(ws, plantilla, month, year, month_name):
    """Clear cached pivot data, write labels and formulas for Ventas LN."""
    # Clear all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Rewrite plantilla lookup (header at row 2, data at row 3 per template)
    _write_plantilla_lookup(ws, plantilla, header_row=2, data_start_row=3)

    # Decorative label
    _cell(ws, 2, 4, f"Ventas por Linea de Negocio {year}",
          font=FONT_BOLD, fill=GREEN_FILL)

    # Page field hints (overwritten by pivot refresh)
    ws.cell(1, 3).value = month

    # "Fecha inc" header
    ws.cell(7, 1).value = "Fecha inc"

    # XLOOKUP formulas for Fecha Incorporacion (col A, data rows 8+)
    for r in range(8, 108):
        ws.cell(r, 1).value = f'=IFERROR(_xlfn.XLOOKUP(B{r},$AF$2:$AF$71,$AJ$2:$AJ$71," "),"")'

    # Extend pivot location for more room
    for pt in ws._pivots:
        pt.location.ref = "B6:E110"


# Cache field indices in the DATA NEW pivot cache
_FLD_MES = 17
_FLD_ANO = 18
_FLD_ANO_INICIO = 20


def _update_pivot_filters(wb, month, year, month_name):
    """Update pivot table filters for the target month and year.

    Updates filter settings on all 5 pivot tables:
    - TablaDinamica1 (Report 1): Mes = target month, Ano = target year
    - TablaDinamica2 (Report 2): Mes = 1..month (YTD), Ano = target year
    - TablaDinamica3 (Report 3): Ano Inicio = target year
    - TablaDinamica4 (Report 4): all months visible on col axis, Ano = target year
    - Ventas LN TablaDinamica1: Mes = 1..month (YTD), Ano = target year

    Sets refreshOnLoad=True on all pivot caches.
    Returns number of pivots updated.
    """
    report_sheet_name = f"Reportes TA {month_name}"
    ytd_months = set(range(1, month + 1))
    all_months = set(range(1, 13))
    updated = 0

    if report_sheet_name in wb.sheetnames:
        ws = wb[report_sheet_name]
        for pt in ws._pivots:
            pt.cache.refreshOnLoad = True

            if pt.name == "TablaDin\u00e1mica1":
                # Report 1: single month filter
                pf_idx = _get_pf_item_for_value(pt, _FLD_MES, month)
                if pf_idx is not None:
                    _set_pf_page_item(pt, _FLD_MES, pf_idx)
                yr_idx = _get_pf_item_for_value(pt, _FLD_ANO, year)
                if yr_idx is not None:
                    _set_pf_page_item(pt, _FLD_ANO, yr_idx)
                updated += 1

            elif pt.name == "TablaDin\u00e1mica2":
                # Report 2: YTD (months 1..month)
                _set_pf_multiselect(pt, _FLD_MES, ytd_months)
                yr_idx = _get_pf_item_for_value(pt, _FLD_ANO, year)
                if yr_idx is not None:
                    _set_pf_page_item(pt, _FLD_ANO, yr_idx)
                updated += 1

            elif pt.name == "TablaDin\u00e1mica3":
                # Report 3: Ano Inicio filter
                yr_idx = _get_pf_item_for_value(pt, _FLD_ANO_INICIO, year)
                if yr_idx is not None:
                    _set_pf_page_item(pt, _FLD_ANO_INICIO, yr_idx)
                updated += 1

            elif pt.name == "TablaDin\u00e1mica4":
                # Report 4: all months visible on column axis
                _set_axis_items_visible(pt, _FLD_MES, all_months)
                yr_idx = _get_pf_item_for_value(pt, _FLD_ANO, year)
                if yr_idx is not None:
                    _set_pf_page_item(pt, _FLD_ANO, yr_idx)
                updated += 1

    if _SECONDARY_SHEET in wb.sheetnames:
        ws = wb[_SECONDARY_SHEET]
        for pt in ws._pivots:
            pt.cache.refreshOnLoad = True
            _set_pf_multiselect(pt, _FLD_MES, ytd_months)
            yr_idx = _get_pf_item_for_value(pt, _FLD_ANO, year)
            if yr_idx is not None:
                _set_pf_page_item(pt, _FLD_ANO, yr_idx)
            updated += 1

    return updated


# ── Force pivot refresh (Win + Mac) ──────────────────────────────────────

def _force_pivot_refresh(wb, n_data_rows):
    """Fuerza el recalculo de pivot tables y formulas al abrir el archivo.

    Mecanismos:
      1. fullCalcOnLoad=True  → recalcula todas las formulas (XLOOKUP, etc.)
      2. refreshOnLoad=True   → pide a Excel que refresque cada pivot cache
      3. recordCount=0        → provoca un mismatch que fuerza rebuild del cache
      4. Rango fuente actualizado → garantiza que el cache cubra todos los datos
      5. Registros cacheados vacios → evita que Windows muestre datos antiguos
    """
    # 1. fullCalcOnLoad para formulas
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True

    # 2-5. Procesar cada pivot cache (sin repetir caches compartidos)
    last_row = n_data_rows + 1  # +1 por header
    seen = set()

    for ws in wb.worksheets:
        for pt in getattr(ws, '_pivots', []):
            cache = pt.cache
            cid = id(cache)
            if cid in seen:
                continue
            seen.add(cid)

            # refreshOnLoad
            cache.refreshOnLoad = True

            # recordCount=0 fuerza a Excel a detectar inconsistencia
            cache.recordCount = 0

            # Actualizar rango fuente del cache para cubrir todos los datos
            try:
                ws_src = cache.source.worksheetSource
                if ws_src and ws_src.ref:
                    parts = ws_src.ref.split(':')
                    if len(parts) == 2:
                        end_col = ''.join(c for c in parts[1] if c.isalpha())
                        ws_src.ref = f"A1:{end_col}{last_row}"
            except (AttributeError, TypeError):
                pass

            # Vaciar registros cacheados para que no haya datos antiguos
            try:
                if cache.records is not None:
                    cache.records.r = []
            except (AttributeError, TypeError):
                pass

    return len(seen)


# ── SUMMARY sheet (fallback para Apple Numbers) ─────────────────────────

def _write_summary_sheet(wb, enriched, year, month, month_name, fecha_inc_lookup):
    """Crea una hoja SUMMARY con todos los reportes pre-calculados en Python.

    Esta hoja muestra datos correctos en CUALQUIER aplicacion (Excel Win,
    Excel Mac, Apple Numbers, Google Sheets) porque los valores son estaticos,
    no dependen de pivot tables.
    """
    if "SUMMARY" in wb.sheetnames:
        del wb["SUMMARY"]
    ws = wb.create_sheet("SUMMARY")

    # Report 1 (mes actual) + Report 2 (YTD) lado a lado
    next_row = _write_report_1_2(ws, enriched, year, month, month_name)

    # Report 3 – Venta por Mes de Inicio
    next_row = _write_report_3(ws, enriched, year, next_row + 2)

    # Report 4 – Venta por Mes (con Fecha Incorporacion)
    next_row = _write_report_4(ws, enriched, year, next_row + 2, fecha_inc_lookup)

    # Anchos de columna
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    for col in range(3, 16):
        ws.column_dimensions[get_column_letter(col)].width = 16


# ── Hoja FX RATES ────────────────────────────────────────────────────────

_FX_CURRENCIES = ["USD", "GBP", "CHF", "JPY", "MXN"]

def _write_fx_rates_sheet(wb):
    """Crea la hoja FX RATES con tasas diarias del mes actual (datos BCE)."""
    if "FX RATES" in wb.sheetnames:
        del wb["FX RATES"]
    ws = wb.create_sheet("FX RATES")

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")

    headers = ["Fecha"]
    for cur in _FX_CURRENCIES:
        headers.append(f"{cur}\u2192EUR")
        headers.append(f"{cur}\u2192USD")
    ncols = len(headers)

    month_label, daily = get_current_month_daily_rates()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1, f"Tipos de cambio BCE \u2014 {month_label}").font = Font(bold=True, size=12)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 14
    for c in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for i, entry in enumerate(daily, 4):
        ws.cell(i, 1, entry["date"]).number_format = "DD/MM/YYYY"
        col = 2
        for cur in _FX_CURRENCIES:
            rates = entry["rates"].get(cur, {})
            cell_eur = ws.cell(i, col, rates.get("EUR", 0))
            cell_usd = ws.cell(i, col + 1, rates.get("USD", 0))
            cell_eur.number_format = '0.000000'
            cell_usd.number_format = '0.000000'
            col += 2

    return month_label, len(daily)


# ── Funcion principal ────────────────────────────────────────────────────

# Hoja del template que contiene los 4 pivots principales + hoja secundaria
_PIVOT_SHEET = "Reportes TA Enero"
_SECONDARY_SHEET = "Ventas por Linea de Negocio"

# Hojas a conservar en el output (ademas de FX RATES que se crea despues)
_KEEP_SHEETS = {_PIVOT_SHEET, _SECONDARY_SHEET, "DATA NEW"}


def _cleanup_sheets(wb, month_name):
    """Elimina hojas innecesarias del template y renombra la principal al mes actual."""
    # Eliminar hojas que no necesitamos
    to_delete = [sn for sn in wb.sheetnames if sn not in _KEEP_SHEETS]
    for sn in to_delete:
        del wb[sn]

    # Renombrar hoja de pivots al mes actual
    new_name = f"Reportes TA {month_name}"
    if _PIVOT_SHEET in wb.sheetnames and _PIVOT_SHEET != new_name:
        wb[_PIVOT_SHEET].title = new_name

    # Ordenar: reporte primero, ventas LN, data, fx rates
    desired = [new_name, _SECONDARY_SHEET, "DATA NEW"]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))


def generate_ta_monthly_report(data_rows, output_dir, fx, year=None, month=None):
    """
    Genera el Reporte Mensual Consolidado de TAs.

    Copia el template, limpia hojas innecesarias, escribe DATA NEW
    con datos enriquecidos, actualiza filtros de tablas dinamicas
    y configura refreshOnLoad para que se actualicen al abrir en Excel.

    Args:
        data_rows: Lista de dicts con claves A-Z (salida de build_data_rows).
        output_dir: Directorio de salida.
        fx: Dict de tipos de cambio {currency: {"EUR": rate, "USD": rate}}.
        year: Ano del reporte (default: ano actual).
        month: Mes del reporte (default: mes actual).

    Returns:
        Ruta del archivo generado.
    """
    today = datetime.now()
    year = year or today.year
    month = month or today.month
    month_name = MONTH_NAMES_ES.get(month, str(month)).capitalize()

    # Verificar template
    if not os.path.exists(TA_TEMPLATE_PATH):
        print(f"  ERROR: No se encuentra el template: {TA_TEMPLATE_PATH}")
        print(f"  Ejecuta 'python3 tools/create_ta_seed.py' y sigue las instrucciones.")
        return None

    # Cargar plantilla y enriquecer datos
    plantilla = load_plantilla_corsario()
    enriched = enrich_data_rows(data_rows, plantilla, fx)

    print(f"  Plantilla Corsario: {len(plantilla)} TAs")
    print(f"  Datos enriquecidos: {len(enriched)} filas")

    # Copiar template
    filename = f"Reporte_TAs_{month_name}_{year}.xlsx"
    filepath = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)
    shutil.copy2(TA_TEMPLATE_PATH, filepath)

    # Abrir y limpiar hojas innecesarias
    wb = load_workbook(filepath)

    if "DATA NEW" not in wb.sheetnames:
        print(f"  ERROR: El template no tiene hoja 'DATA NEW'")
        wb.close()
        return None

    _cleanup_sheets(wb, month_name)
    print(f"  Hojas: {wb.sheetnames}")

    # Escribir datos enriquecidos en DATA NEW
    ws = wb["DATA NEW"]
    _write_data_new(ws, enriched, plantilla, fx)
    print(f"  DATA NEW: {len(enriched)} filas escritas")

    # Hoja FX RATES
    month_label, n_days = _write_fx_rates_sheet(wb)
    print(f"  Hoja FX RATES: {n_days} dias de {month_label}")

    # Preparar hojas de reportes (limpiar cache, ajustar layout, labels)
    report_sheet_name = f"Reportes TA {month_name}"
    ws_report = wb[report_sheet_name]
    ws_ventas = wb[_SECONDARY_SHEET]
    _prepare_report_sheet(ws_report, plantilla, month, year, month_name)
    _prepare_ventas_sheet(ws_ventas, plantilla, month, year, month_name)
    print(f"  Layout ajustado, labels y formulas escritos")

    # Actualizar filtros de tablas dinamicas
    n_updated = _update_pivot_filters(wb, month, year, month_name)
    print(f"  Pivot tables actualizados: {n_updated} (refreshOnLoad=True)")

    # Forzar recalculo al abrir (pivots + formulas) — fix para Excel Windows
    n_caches = _force_pivot_refresh(wb, len(enriched))
    print(f"  Force refresh: {n_caches} caches (fullCalcOnLoad + records cleared)")

    # Hoja SUMMARY con datos pre-calculados (compatible con Apple Numbers)
    fecha_inc_lookup = _build_fecha_inc_lookup(enriched)
    _write_summary_sheet(wb, enriched, year, month, month_name, fecha_inc_lookup)
    print(f"  Hoja SUMMARY creada (datos pre-calculados, compatible Numbers)")

    # Ordenar hojas: reportes → SUMMARY → DATA NEW → FX RATES
    desired_order = [
        f"Reportes TA {month_name}", _SECONDARY_SHEET,
        "SUMMARY", "DATA NEW", "FX RATES",
    ]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    # Guardar
    wb.save(filepath)
    wb.close()

    return filepath
