#!/usr/bin/env python3
"""
Genera el archivo semilla (seed) para crear el template del Reporte de TAs.

Uso:
    python3 tools/create_ta_seed.py

Genera output/TA_seed.xlsx con:
  - Hoja "DATA NEW": headers + datos enriquecidos (30 columnas)
  - Formulas XLOOKUP para conversion FX (cols L, M, O, P, Q, AB, AC)
  - FX lookup table en cols AO-AQ
  - Plantilla Corsario en cols AI-AM

Despues de ejecutar:
  1. Abrir TA_seed.xlsx en Microsoft Excel
  2. Crear pivot tables siguiendo las instrucciones en consola
  3. Guardar como templates/Reporte_TAs_template.xlsx
"""

import os
import sys

# Agregar el directorio raiz al path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import ENTITIES, BASE_DIR
from src.data_loader import load_reserva, load_dreserva, compute_rentabilidad
from src.fx_rates import get_latest_fx, preload_fx_months
from src.weekly.data_sheet import build_data_rows
from src.ta_monthly.report import (
    load_plantilla_corsario, enrich_data_rows, _write_plantilla_block,
    _write_fx_table, _get_formula, _cell,
    FONT, FONT_HDR, DATA_NEW_HDR, HDR_BORDER, THIN_BORDER,
    DATE_FMT, PCT_FMT, NUM_FMT, DATA_NEW_COLS, DATA_NEW_WIDTHS,
    FORMULA_KEYS,
)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


def main():
    print("=== Generando archivo semilla para template de TAs ===\n")

    # Obtener FX rates
    print("  Obteniendo tipos de cambio...")
    fx = get_latest_fx()

    # Cargar datos de ambas entidades
    all_data_rows = []
    for key, cfg in ENTITIES.items():
        data_dir = cfg["data_dir"]
        if not os.path.exists(data_dir):
            print(f"  Saltando {cfg['label']}: {data_dir} no existe")
            continue

        print(f"  Cargando {cfg['label']}...")
        reserva_df = load_reserva(data_dir)
        dreserva_df = load_dreserva(data_dir)
        rentabilidad_df = compute_rentabilidad(dreserva_df)
        preload_fx_months(reserva_df["fecha"].dropna())
        data_rows = build_data_rows(reserva_df, rentabilidad_df, cfg["company"])
        all_data_rows.extend(data_rows)
        print(f"    {len(data_rows)} filas")

    print(f"\n  Total: {len(all_data_rows)} filas")

    # Enriquecer
    plantilla = load_plantilla_corsario()
    enriched = enrich_data_rows(all_data_rows, plantilla)
    print(f"  Plantilla Corsario: {len(plantilla)} TAs")
    print(f"  Datos enriquecidos: {len(enriched)} filas")

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA NEW"

    # Headers
    for ci, (_, header) in enumerate(DATA_NEW_COLS, 1):
        _cell(ws, 1, ci, header, font=FONT_HDR, fill=DATA_NEW_HDR, border=HDR_BORDER)

    # Datos con formulas XLOOKUP
    for ri, row_data in enumerate(enriched, 2):
        for ci, (key, _) in enumerate(DATA_NEW_COLS, 1):
            if key in FORMULA_KEYS:
                formula = _get_formula(key, ri)
                cell = ws.cell(ri, ci, formula)
                cell.font = FONT
                cell.border = THIN_BORDER
                if key == "Q":
                    cell.number_format = PCT_FMT
                else:
                    cell.number_format = NUM_FMT
            else:
                val = row_data.get(key)
                if val is not None:
                    fmt = None
                    if key in ("D", "E", "F", "V", "AD_fecha_inc"):
                        if isinstance(val, datetime):
                            fmt = DATE_FMT
                    elif key in ("J", "N"):
                        fmt = NUM_FMT
                    elif key == "AA_com":
                        fmt = PCT_FMT
                    _cell(ws, ri, ci, val, fmt=fmt, border=THIN_BORDER)
                else:
                    _cell(ws, ri, ci, None, border=THIN_BORDER)

    # Anchos
    for i, w in enumerate(DATA_NEW_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Plantilla Corsario en cols AI-AM (35-39)
    _write_plantilla_block(ws, plantilla, col_start=35, row_start=1)

    # FX lookup table en cols AO-AQ (41-43)
    _write_fx_table(ws, fx)

    # Guardar
    output_dir = os.path.join(BASE_DIR, "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "TA_seed.xlsx")
    wb.save(output_path)
    wb.close()

    print(f"\n  Archivo semilla generado: {output_path}")
    print(f"  DATA NEW: {len(enriched)} filas x {len(DATA_NEW_COLS)} columnas")
    print(f"  Formulas XLOOKUP: L, M, O, P, Q, AB, AC")
    print(f"  FX lookup table: cols AO-AQ")

    # Instrucciones
    print(f"""
{'=' * 70}
  INSTRUCCIONES PARA CREAR EL TEMPLATE
{'=' * 70}

  1. Abrir {output_path} en Microsoft Excel

  2. Crear una nueva hoja "Reportes TA" y crear 4 pivot tables:

     REPORT 1 - "Venta mes":
       Source: DATA NEW (A1:AD{len(enriched)+1})
       Filtros: Compania, Ano, Mes
       Filas: Oficina > Linea de Negocio > vendedor
       Valores: Sum(Total Venta USD), Sum(Renta after COM USD),
                Sum(Rentabilidad en USD)

     REPORT 2 - "Venta acumulada":
       Source: DATA NEW
       Filtros: Compania, Ano
       Filas: Oficina > Linea de Negocio > vendedor
       Valores: Sum(Total Venta USD), Sum(Renta after COM USD),
                Sum(Rentabilidad en USD)

     REPORT 3 - "Venta Mes de Inicio":
       Source: DATA NEW
       Filtros: Compania, Ano Inicio, Oficina, Linea de Negocio
       Filas: vendedor
       Columnas: Mes Inicio
       Valores: Sum(Total Venta USD)

     REPORT 4 - "Venta por Mes":
       Source: DATA NEW
       Filtros: Compania, Ano
       Filas: Fecha Incorporacion, vendedor
       Columnas: Mes
       Valores: Sum(Total Venta USD)

  3. Crear una nueva hoja "Ventas por Linea de Negocio" con 1 pivot:
       Source: DATA NEW
       Filtros: Compania, Ano, Mes, Ano Inicio
       Filas: Oficina > Linea de Negocio > vendedor
       Valores: Sum(Total Venta USD), Sum(Renta after COM USD),
                Sum(Rentabilidad en USD)

  4. Guardar como: templates/Reporte_TAs_template.xlsx

  5. Listo! El script main.py usara este template automaticamente.
{'=' * 70}
""")


if __name__ == "__main__":
    main()
